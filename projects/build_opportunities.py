#!/usr/bin/env python3
"""Build the canonical FairBanks opportunity tracker Excel.

Writes projects/opportunities.xlsx per projects/.cursor/rules/source_of_truth.mdc.
Curated from deep scans of Opportunities for Youth (and future target sites).
Only Uganda-eligible, still-open opportunities; gender-tagged.
"""

from datetime import datetime
from pathlib import Path
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent
DEFAULT_OUT = ROOT / "opportunities.xlsx"

GENDER_BASED = "Gender-based (women/girls)"
MULTI_GENDER = "Multi-gender (all genders)"

OPPORTUNITIES = [
    {
        "title": "AWIEF Pitch n Grow 2026 — Women-Led Deep-Tech Startup Competition",
        "url": "https://opportunitiesforyouth.org/2026/06/25/awief-pitch-n-grow-2026-applications-open-for-africas-leading-women-led-deep-technology-startup-competition/",
        "gender": GENDER_BASED,
        "description": (
            "Continent-wide competition for women-led ventures using deep "
            "technology under the theme 'Deep Roots. Digital Futures.' Tracks: "
            "Idea, Startup, Scale-Up. HealthTech (digital healthcare, "
            "telemedicine, AI diagnostics) and AI/ML are priority sectors, so "
            "FCHIP is a strong fit. Benefits include a training bootcamp, "
            "mentorship, pitch coaching, investor visibility, media exposure, "
            "travel support for finalists, and cash prizes. Open to women "
            "founders across Africa, so Ugandans qualify. Submit before "
            "screening begins."
        ),
        "deadline": "20 July 2026",
        "deadline_sort": "2026-07-20",
    },
    {
        "title": "The Gadfly Project Custom Web Application Grant 2026",
        "url": "https://opportunitiesforyouth.org/2026/07/16/the-gadfly-project-custom-web-application-grant-2026-apply-for-in-kind-software-development-support-worth-up-to-us100000/",
        "gender": MULTI_GENDER,
        "description": (
            "In-kind grant giving mission-driven organisations professional "
            "software development (valued US$5,000–100,000) instead of cash. "
            "Gadfly's in-house team helps you build custom web apps, databases, "
            "outreach platforms, or internal systems that expand community "
            "impact. Open to nonprofits, charities, social enterprises, and "
            "community organisations worldwide, so a Ugandan health "
            "organisation qualifies. Reviewed quarterly; selection weighs "
            "users served and community impact over technical complexity. "
            "Useful for building FCHIP mobile and data tools."
        ),
        "deadline": "24 July 2026 (next cycle 23 October 2026)",
        "deadline_sort": "2026-07-24",
    },
    {
        "title": "AUC Venture Lab Women Innovation Fellowship 2026",
        "url": "https://opportunitiesforyouth.org/2026/07/16/auc-venture-lab-women-innovation-fellowship-empowering-women-entrepreneurs-in-egypt-and-beyond/",
        "gender": GENDER_BASED,
        "description": (
            "American University in Cairo fellowship (1 Sep–31 Oct 2026, "
            "Cairo) empowering women entrepreneurs and innovators through "
            "leadership training, ecosystem access, mentorship, and "
            "cross-regional exposure across Egypt and Denmark. Open to women "
            "across sectors including HealthTech, AI, and sustainability, and "
            "to founders beyond Egypt, so Ugandan women innovators may apply. "
            "Strong networking and investment-readiness support for a "
            "woman-led health venture."
        ),
        "deadline": "25 July 2026",
        "deadline_sort": "2026-07-25",
    },
    {
        "title": "Young Feminist AI School 2026 (UN Women)",
        "url": "https://opportunitiesforyouth.org/2026/07/16/ai-for-gender-equality-un-women-ai-school-opens-for-changemakers/",
        "gender": GENDER_BASED,
        "description": (
            "Free, fully online 12-week UN Women programme (Aug–Oct 2026) "
            "teaching practical, no-code AI skills for advocacy, policy, civic "
            "tech, and responsible AI. Open to young people aged 18–30 "
            "worldwide, with at least 70% of places reserved for young women "
            "and girls. About four hours weekly; every participant builds a "
            "real project. Good capacity building for team members applying "
            "gender-responsive AI to community health. Ugandans eligible."
        ),
        "deadline": "28 July 2026",
        "deadline_sort": "2026-07-28",
    },
    {
        "title": "GirlCode Hackathon 2026 (includes Kampala, Uganda)",
        "url": "https://opportunitiesforyouth.org/2026/07/15/building-africas-next-generation-of-ai-powered-female-innovators-apply-for-the-2026-girlcode-hackathon-across-six-african-countries/",
        "gender": GENDER_BASED,
        "description": (
            "Multi-country AI hackathon for female innovators across South "
            "Africa, Kenya, Uganda, Tanzania, and Ghana, building Africa's "
            "next generation of women in tech. The Kampala event runs 5–6 "
            "September 2026 at UICT Nakawa. Great for visibility, "
            "team-building, recruiting women technologists, or showcasing "
            "FCHIP concepts. No single formal deadline is stated, so register "
            "before the local event date. Ugandan women eligible."
        ),
        "deadline": "Not formally stated — register before Kampala event 5–6 Sep 2026",
        "deadline_sort": "2026-09-05",
    },
    {
        "title": "U.S. Dept of State — Uganda Health System MOU (up to $60M)",
        "url": "https://opportunitiesforyouth.org/2026/06/15/u-s-department-of-state-launches-up-to-60-million-funding-opportunity-to-strengthen-ugandas-health-system-through-the-health-foreign-assistance-mou-implementation-plan/",
        "gender": MULTI_GENDER,
        "description": (
            "Major U.S. Department of State call (DFOP0017890) to strengthen "
            "Uganda's health system under the U.S.–Uganda Health MOU: up to "
            "$60M across ~15 awards over five years. Objective 4 funds scaling "
            "digital health systems (EMRs, DHIS2, eCHIS, iHRIS, national data "
            "warehouse); other tracks cover community-based health financing, "
            "faith/community facilities, and supply-chain accountability. "
            "Two-step process starting with a Statement of Interest. Highly "
            "aligned with FCHIP and FairBanks community health work. Ugandan "
            "organisations are the intended applicants."
        ),
        "deadline": "31 July 2026 (Statement of Interest, 5:00 PM EDT)",
        "deadline_sort": "2026-07-31",
    },
    {
        "title": "World Health Summit 2026 Youth Group Opportunity (Berlin)",
        "url": "https://opportunitiesforyouth.org/2026/06/26/apply-now-world-health-summit-2025-stipend-program-for-global-health-changemakers/",
        "gender": MULTI_GENDER,
        "description": (
            "Complimentary group tickets (3–15 per organisation) for "
            "youth-led organisations to attend all three days of the World "
            "Health Summit 2026 in Berlin. Open to youth-led groups working "
            "in health, advocacy, education, or social impact, with nominees "
            "aged 28 or under; travel, visa, and accommodation are "
            "self-funded. Diverse and underrepresented regions are "
            "encouraged, so a Ugandan community-health organisation is a "
            "strong candidate for global visibility and networking."
        ),
        "deadline": "31 July 2026",
        "deadline_sort": "2026-07-31",
    },
    {
        "title": "African Union Youth SRH Ambassadors Initiative 2026 (AUDA-NEPAD)",
        "url": "https://opportunitiesforyouth.org/2026/06/13/african-union-youth-srh-ambassadors-initiative-2026-apply-to-become-an-auda-nepad-youth-srh-ambassador/",
        "gender": MULTI_GENDER,
        "description": (
            "African Union (AUDA-NEPAD) initiative recruiting young Africans "
            "aged 18–35 as voluntary Sexual and Reproductive Health "
            "ambassadors for advocacy, awareness, and community mobilisation. "
            "Offers recognition, training, mentorship, networking, and "
            "sponsored participation in selected events (no salary). "
            "Applicants should show SRH commitment and an active digital "
            "presence; diverse genders and regions are encouraged. Relevant "
            "to FairBanks community reach and youth health engagement. "
            "Ugandans eligible — note the deadline is today."
        ),
        "deadline": "17 July 2026 (closing)",
        "deadline_sort": "2026-07-17",
    },
    {
        "title": "Brandtech Group AI for Good Scholarship 2026 (One Young World)",
        "url": "https://opportunitiesforyouth.org/2026/07/15/ai-for-good-scholarship-2025-empowering-young-innovators-to-address-global-challenges/",
        "gender": MULTI_GENDER,
        "description": (
            "Fully funded scholarship for five young African leaders (aged "
            "18–30) using AI for social, health, education, environmental, or "
            "civic impact, to attend the One Young World Summit 2026 in Cape "
            "Town. Healthcare is a named priority area. Requires African "
            "nationality and residency plus a proven record of responsible AI "
            "work. A strong platform for a founder or team member advancing AI "
            "in community health. Ugandans eligible."
        ),
        "deadline": "2 August 2026",
        "deadline_sort": "2026-08-02",
    },
    {
        "title": "Africa CDC African Epidemic Services (AES) Epidemiology Fellowship 2026–2028",
        "url": "https://opportunitiesforyouth.org/2026/07/16/africa-cdc-african-epidemic-services-aes-epidemiology-fellowship-2026-2028-fully-funded-apply-by-26-august-2026/",
        "gender": MULTI_GENDER,
        "description": (
            "Fully funded two-year Africa CDC fellowship — three months' "
            "training in Addis Ababa plus 21 months' field placement — for "
            "citizens of African Union member states under 35 already employed "
            "in public health, with a relevant degree and three years' "
            "experience. Builds outbreak investigation, surveillance, and "
            "data-analysis skills; a Public Health Informatics track also "
            "exists. Covers stipend, travel, insurance, hardware, and "
            "software. Excellent professional development for FairBanks "
            "epidemiology or informatics talent. Ugandans eligible."
        ),
        "deadline": "26 August 2026",
        "deadline_sort": "2026-08-26",
    },
    {
        "title": "IFAD Grant Call — Rural Sector Performance Assessment (up to US$1.2M)",
        "url": "https://opportunitiesforyouth.org/2026/07/14/ifad-announces-us1-2-million-grant-call-for-proposals-to-strengthen-rural-development-policies-worldwide/",
        "gender": MULTI_GENDER,
        "description": (
            "IFAD call for proposals worth up to US$1.2M to implement "
            "'Leveraging IFAD's Rural Sector Performance Assessment for Policy "
            "and Investment.' Open to non-profits, private companies, "
            "universities, think tanks, and research institutes (20–25% "
            "counterpart contribution). Optional non-binding Expression of "
            "Interest by 31 July 2026. Best fit if FairBanks partners on rural "
            "policy and data with livelihood–health linkages; verify thematic "
            "match before investing proposal effort. Ugandan organisations "
            "eligible."
        ),
        "deadline": "4 September 2026 (12:00 CET); optional EOI 31 July 2026",
        "deadline_sort": "2026-09-04",
    },
    {
        "title": "World Bank GovTech Innovation Challenge 2026",
        "url": "https://opportunitiesforyouth.org/2026/07/10/world-bank-govtech-innovation-challenge-2026-global-call-for-proposals-to-build-digital-solutions-for-governments/",
        "gender": MULTI_GENDER,
        "description": (
            "Global World Bank / SECO / Trust Valley challenge connecting "
            "technology companies with government digitalisation needs. "
            "Selected innovators get fully funded capacity building, a "
            "Switzerland bootcamp, and a chance to build a proof of concept "
            "with a government partner. Current open challenges are Moroccan "
            "public-finance audit use cases, but new calls appear over time — "
            "watch for health or Africa challenges matching district health "
            "intelligence. Deadlines vary by challenge. Companies worldwide, "
            "including Ugandan, may apply."
        ),
        "deadline": "Varies by challenge — check the specific challenge page",
        "deadline_sort": "9999-12-31",
    },
]


def word_count(text: str) -> int:
    return len(text.split())


def main() -> None:
    for row in OPPORTUNITIES:
        wc = word_count(row["description"])
        if wc > 100:
            raise SystemExit(f"Description too long ({wc}): {row['title']}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    headers = [
        "Project Title",
        "URL",
        "Brief Description",
        "Application Deadline",
        "Gender Category",
    ]
    header_fill = PatternFill("solid", fgColor="0B3D2E")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    alt_fill = PatternFill("solid", fgColor="F3F8F5")
    gender_fill = PatternFill("solid", fgColor="F6E7F2")
    multi_fill = PatternFill("solid", fgColor="E7F0FA")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")

    # Group by gender category, then soonest deadline within each group.
    rows = sorted(
        OPPORTUNITIES,
        key=lambda r: (0 if r["gender"] == GENDER_BASED else 1, r["deadline_sort"], r["title"]),
    )

    for i, row in enumerate(rows, 2):
        values = [row["title"], row["url"], row["description"], row["deadline"], row["gender"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i, col, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if i % 2 == 0:
                cell.fill = alt_fill
        # Colour-code the gender category cell so the two groups are obvious.
        gcell = ws.cell(i, 5)
        gcell.fill = gender_fill if row["gender"] == GENDER_BASED else multi_fill
        gcell.font = Font(name="Calibri", size=10, bold=True)
        ws.row_dimensions[i].height = 92

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 74
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 22
    ws.row_dimensions[1].height = 26
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"
    ws.freeze_panes = "A2"

    # Scan notes / methodology sheet.
    meta = wb.create_sheet("Scan Notes")
    gb = sum(1 for r in rows if r["gender"] == GENDER_BASED)
    mg = len(rows) - gb
    notes = [
        ("Source site", "https://opportunitiesforyouth.org/"),
        ("Scanned on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Applicant context", "FairBanks / FCHIP — a Ugandan community health intelligence venture. We are Ugandans."),
        ("Method", "Live deep scan of the homepage and category sections (Grants, Fellowship, Jobs, Conferences, Internships, Volunteering, Competition) plus each individual listing page, verified one by one."),
        ("Eligibility filter", "Included ONLY opportunities that are (a) still open as of the scan date and (b) open to Ugandan applicants (global, pan-African / African Union, or Uganda-specific). Country-restricted calls that exclude Uganda were dropped."),
        ("Excluded (country)", "Pathways to Scale (Ethiopia/Ghana/Nigeria/Rwanda only), SPARK Change-Makers (Ghana), ADB dengue/laundry challenges (Asia-Pacific pilots), Supercell grants (game-dev studios only)."),
        ("Excluded (closed)", "Moonshot Awards (12 Jul), WHO Youth Council (30 Jun), AIMS SDG Challenge (1 Jun), A+ Alliance Gender & AI Cohort 2 (25 Apr) — deadlines already passed."),
        ("Gender categories", f"Gender-based (women/girls): {gb}. Multi-gender (all genders): {mg}. Total: {len(rows)}."),
        ("Gender-based list", "AWIEF Pitch n Grow, AUC Venture Lab Women Innovation Fellowship, Young Feminist AI School (UN Women), GirlCode Hackathon."),
        ("Priority soon", "AUDA-NEPAD SRH Ambassadors (17 Jul, today), AWIEF (20 Jul), Gadfly grant (24 Jul), AUC Women Fellowship (25 Jul), Young Feminist AI School (28 Jul), US State Dept Uganda SOI / WHS Youth Group (31 Jul)."),
        ("Best fit for FairBanks", "US State Dept Uganda Health MOU, AWIEF Pitch n Grow, WHS Youth Group, Africa CDC AES Fellowship, Gadfly Web App Grant."),
        ("Important", "Deadlines and eligibility are quoted from Opportunities for Youth article text. ALWAYS confirm on the official programme page before applying — details can change."),
    ]
    meta["A1"] = "Field"
    meta["B1"] = "Detail"
    meta["A1"].font = Font(bold=True, color="FFFFFF")
    meta["B1"].font = Font(bold=True, color="FFFFFF")
    meta["A1"].fill = header_fill
    meta["B1"].fill = header_fill
    for i, (k, v) in enumerate(notes, 2):
        meta.cell(i, 1, k).font = Font(bold=True)
        meta.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
        meta.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
        meta.row_dimensions[i].height = 42
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 112

    out = Path(os.environ.get("OPPORTUNITIES_OUT", str(DEFAULT_OUT)))
    wb.save(out)
    print(f"Wrote {len(rows)} opportunities to {out}")
    print(f"Gender-based: {gb} | Multi-gender: {mg}")
    for r in rows:
        print(f"- [{r['gender'][:12]:12}] {r['deadline'][:34]:34} | {r['title'][:60]}")


if __name__ == "__main__":
    main()

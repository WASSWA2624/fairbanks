#!/usr/bin/env python3
"""Scan Simpler.Grants.gov for FairBanks-eligible open grants.

Writes opportunities/new_grants.xlsx — soonest deadline first.

Source catalog: Simpler.Grants.gov (same federal opportunities as Grants.gov).
Search uses the public Grants.gov search2 + fetchOpportunity APIs (no key),
resolves each row to a Simpler.Grants.gov opportunity URL, and optionally
uses SIMPLER_GRANTS_API_KEY when set.

Applicant context (from .cursor/source-of-truth Blueprint + opportunities SoT):
FAIRBANKS MEDICAL CENTRE LIMITED — women-led Ugandan private company
(not NGO); Community Reach + FCHIP (community / digital / climate health).
"""

from __future__ import annotations

import os
import re
import time
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Any
from urllib.parse import quote

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "new_grants.xlsx"

SIMPLER_BASE = "https://simpler.grants.gov"
SIMPLER_API = "https://api.simpler.grants.gov"
SEARCH2 = "https://api.grants.gov/v1/api/search2"
FETCH = "https://api.grants.gov/v1/api/fetchOpportunity"

GENDER_BASED = "Gender-based (women/girls)"
MULTI_GENDER = "Multi-gender (all genders)"

# FairBanks as a private company — prefer these eligibility codes.
ELIG_FOR_PROFIT = {"22", "23", "99"}  # for-profit, small business, unrestricted
ELIG_FILTER = "22|23|99"

HEADERS = {
    "User-Agent": "FairBanksOpportunityScan/1.0 (+https://simpler.grants.gov)",
    "Content-Type": "application/json",
    "Accept": "application/json",
}

# Queries aimed at FairBanks / FCHIP themes + Uganda / Africa geography.
SEARCH_QUERIES = [
    "Uganda",
    "East Africa health",
    "global health security",
    "partner countries outbreak",
    "community health Africa",
    "disease surveillance Africa",
    "maternal health Africa",
    "HIV prevention Africa",
    "CDC-RFA-JG",
    "DHAPP",
    "Advancing Global Health",
]

# Strong geography for a Ugandan applicant (avoid loose "global"/"international").
GEO_STRONG = re.compile(
    r"\b("
    r"uganda|kampala|east africa|eastern africa|sub[- ]?saharan africa|"
    r"partner countr(?:y|ies)|african countr|"
    r"low[- ] and middle[- ]income countr|lmics?"
    r")\b",
    re.I,
)
GEO_AFRICA = re.compile(r"\b(africa|african)\b", re.I)
# Country-specific RFAs that are NOT Uganda.
OTHER_COUNTRY_LOCK = re.compile(
    r"\b("
    r"kenya|kenyan|democratic republic of (?:the )?congo|\bdrc\b|"
    r"c(?:[!&]|ô|&ocirc;)te d['\"]?ivoire|ivory coast|cote d['\"]?ivoire|"
    r"nigeria|ethiopia|tanzania|rwanda|ghana|senegal|malawi|zambia|"
    r"mozambique|botswana|namibia|south africa|cameroon|madagascar|"
    r"sierra leone|liberia|guinea|mali|niger|chad|sudan|south sudan|"
    r"india|indonesia|philippines|vietnam|thailand|cambodia|"
    r"bangladesh|pakistan|nepal|brazil|colombia|mexico|peru|haiti|"
    r"guatemala|honduras|ukraine|georgia|kazakhstan|central america"
    r")\b",
    re.I,
)
FOREIGN_APPLICANT_OK = re.compile(
    r"non[- ]?domestic \(non[- ]?u\.?s\.?\) entities? \(foreign organizations?\) are eligible to apply"
    r"|foreign organizations? are eligible to apply"
    r"|non[- ]?u\.?s\.? entities? are eligible to apply",
    re.I,
)
FOREIGN_COLLAB_ONLY = re.compile(
    r"foreign components? may be included"
    r"|eligible for foreign components?"
    r"|applications may include foreign components?",
    re.I,
)

# Clear US-domestic-only signals.
US_ONLY = re.compile(
    r"\b("
    r"united states only|u\.?s\.? only|domestic only|"
    r"within the united states|states and territories|"
    r"delta states|rural health clinic|hrsa state|"
    r"federally qualified health center|tribal epidemiology|"
    r"indian health service|u\.?s\.? territories|"
    r"alaska native|native hawaiian serving"
    r")\b",
    re.I,
)

# CDC / State / DoD global health programme agencies often open to local partners.
GLOBAL_HEALTH_AGENCIES = {
    "HHS-CDC-GHC",
    "HHS-CDC-NCEZID",
    "HHS-CDC-CGH",
    "DOS-GHSD",
    "DOS-SGHC",
    "DOD-AMRAA",
}

# Thematic fit with FairBanks Blueprint / FCHIP / Community Reach.
THEME_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("surveillance / GHS", re.compile(r"surveillance|outbreak|global health security|7-1-7|detect,? notify,? and respond", re.I)),
    ("community health", re.compile(r"community health|chw|vht|village health|primary health|family(?: and)? community health", re.I)),
    ("maternal / child", re.compile(r"maternal|neonatal|child health|immuni[sz]ation", re.I)),
    ("digital / AI health", re.compile(r"digital health|e[- ]?health|telemedicine|predictive|machine learning|HMIS|EMR|community health intelligence|informatics", re.I)),
    ("climate–health", re.compile(r"climate[- ]health|climate[- ]driven|climate[- ]sensitive|rainfall", re.I)),
    ("HIV / infectious", re.compile(r"\bHIV\b|AIDS|malaria|tuberculosis|\bTB\b|fungal diseas|infectious disease|zoonotic", re.I)),
    ("health systems", re.compile(r"health system strengthening|public health capacity|health security", re.I)),
]

WOMEN_LED = re.compile(r"\bwomen[- ]led\b|\bwomen and girls\b|\bgender equality\b|\bwomen'?s health\b", re.I)


def session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def parse_us_date(value: str | None) -> str:
    """Return YYYY-MM-DD for sorting, or 9999-12-31 if unknown/rolling."""
    if not value or not str(value).strip():
        return "9999-12-31"
    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(text[:10] if fmt == "%Y-%m-%d" else text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # e.g. "Aug 14, 2026 12:00:00 AM EDT"
    m = re.search(r"([A-Za-z]{3,9}\s+\d{1,2},\s+\d{4})", text)
    if m:
        for fmt in ("%b %d, %Y", "%B %d, %Y"):
            try:
                return datetime.strptime(m.group(1), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return "9999-12-31"


def format_deadline(close_date: str | None, response_date_desc: str | None = None) -> str:
    sort = parse_us_date(close_date)
    if sort == "9999-12-31":
        return "Rolling / not listed — verify on Simpler.Grants.gov"
    try:
        pretty = datetime.strptime(sort, "%Y-%m-%d").strftime("%d %B %Y")
    except ValueError:
        pretty = close_date or "Unknown"
    if response_date_desc and "11:59" in response_date_desc:
        return f"{pretty} (11:59 p.m. ET - verify)"
    return pretty


def money(value: Any) -> str:
    if value is None or value == "" or str(value) in {"0", "0.0"}:
        return "Not listed"
    try:
        return f"USD {int(float(str(value).replace(',', ''))):,}"
    except (TypeError, ValueError):
        return str(value)


def word_count(text: str) -> int:
    return len(text.split())


def truncate_words(text: str, limit: int = 100) -> str:
    words = text.split()
    if len(words) <= limit:
        return text
    return " ".join(words[: limit - 1]) + "…"


def search_grantsgov(sess: requests.Session, query: str, start: int = 0, rows: int = 25) -> list[dict[str, Any]]:
    payload = {
        "keyword": query,
        "oppStatuses": "posted",
        "fundingCategories": "HL",
        "eligibilities": ELIG_FILTER,
        "rows": rows,
        "startRecordNum": start,
        "sortBy": "closeDate|asc",
    }
    r = sess.post(SEARCH2, json=payload, timeout=90)
    r.raise_for_status()
    body = r.json()
    if body.get("errorcode") not in (0, "0", None):
        raise RuntimeError(f"search2 error: {body.get('msg')}")
    data = body.get("data") or {}
    return list(data.get("oppHits") or [])


def search_simpler_api(sess: requests.Session, query: str, page: int = 1) -> list[dict[str, Any]]:
    """Optional path when SIMPLER_GRANTS_API_KEY is set."""
    key = os.environ.get("SIMPLER_GRANTS_API_KEY", "").strip()
    if not key:
        return []
    headers = {**HEADERS, "X-API-Key": key}
    payload = {
        "query": query[:100],
        "filters": {
            "opportunity_status": {"one_of": ["posted"]},
            "funding_category": {"one_of": ["health"]},
            "applicant_type": {
                "one_of": [
                    "for_profit_organizations_other_than_small_businesses",
                    "small_businesses",
                    "unrestricted",
                ]
            },
        },
        "pagination": {
            "page_offset": page,
            "page_size": 25,
            "sort_order": [{"order_by": "close_date", "sort_direction": "ascending"}],
        },
    }
    r = sess.post(f"{SIMPLER_API}/v1/opportunities/search", headers=headers, json=payload, timeout=90)
    if r.status_code != 200:
        return []
    return list((r.json().get("data") or []))


def fetch_opportunity(sess: requests.Session, opportunity_id: str | int, tries: int = 3) -> dict[str, Any] | None:
    last_err: Exception | None = None
    for attempt in range(tries):
        try:
            r = sess.post(FETCH, json={"opportunityId": str(opportunity_id)}, timeout=90)
            r.raise_for_status()
            body = r.json()
            data = body.get("data")
            if not isinstance(data, dict) or "opportunityNumber" not in data:
                return None
            return data
        except requests.RequestException as exc:
            last_err = exc
            time.sleep(1.2 * (attempt + 1))
    if last_err:
        raise last_err
    return None


def resolve_simpler_url(sess: requests.Session, opp_number: str, cache: dict[str, str]) -> str:
    if opp_number in cache:
        return cache[opp_number]
    url = f"{SIMPLER_BASE}/search?query={quote(opp_number)}"
    try:
        r = sess.get(url, headers={"User-Agent": HEADERS["User-Agent"], "Accept": "text/html"}, timeout=90)
        r.raise_for_status()
        uuids = re.findall(r"/opportunity/([0-9a-f-]{36})", r.text, flags=re.I)
        if uuids:
            link = f"{SIMPLER_BASE}/opportunity/{uuids[0]}"
            cache[opp_number] = link
            return link
    except requests.RequestException:
        pass
    # Fallback: Simpler search for the opportunity number.
    cache[opp_number] = f"{SIMPLER_BASE}/search?query={quote(opp_number)}"
    return cache[opp_number]


def theme_hits(text: str) -> list[str]:
    return [label for label, pat in THEME_PATTERNS if pat.search(text)]


def is_uganda_eligible(
    text: str,
    applicant_types: list[dict[str, Any]],
    agency_code: str = "",
) -> bool:
    """Ugandan private company must have a real apply path — not US-only NIH."""
    has_uganda = bool(re.search(r"\buganda\b|\bkampala\b", text, re.I))
    locked_elsewhere = bool(OTHER_COUNTRY_LOCK.search(text)) and not has_uganda
    if locked_elsewhere:
        return False
    if US_ONLY.search(text) and not GEO_STRONG.search(text):
        return False
    if GEO_STRONG.search(text) or has_uganda:
        return True
    if GEO_AFRICA.search(text) and not FOREIGN_COLLAB_ONLY.search(text):
        return True
    ids = {str(t.get("id")) for t in applicant_types}
    # CDC GHC / State GHSD / DHAPP-style global health packages with for-profit path.
    if agency_code in GLOBAL_HEALTH_AGENCIES and (ids & ELIG_FOR_PROFIT):
        if re.search(r"global health|outbreak|surveillance|HIV|AIDS|partner", text, re.I):
            return True
    if FOREIGN_APPLICANT_OK.search(text) and not FOREIGN_COLLAB_ONLY.search(text):
        return True
    if "99" in ids and agency_code in GLOBAL_HEALTH_AGENCIES:
        return True
    return False


def entity_ok(applicant_types: list[dict[str, Any]]) -> bool:
    ids = {str(t.get("id")) for t in applicant_types}
    return bool(ids & ELIG_FOR_PROFIT) or not ids  # empty → verify on listing


def build_description(detail: dict[str, Any], themes: list[str], simpler_url: str) -> str:
    syn = detail.get("synopsis") or {}
    title = detail.get("opportunityTitle") or ""
    number = detail.get("opportunityNumber") or ""
    agency = (syn.get("agencyName") or detail.get("owningAgencyCode") or "").strip()
    synopsis = unescape(re.sub(r"\s+", " ", (syn.get("synopsisDesc") or "").strip()))
    types = ", ".join(
        t.get("description", "") for t in (syn.get("applicantTypes") or []) if str(t.get("id")) in ELIG_FOR_PROFIT
    ) or "see listing"
    ceiling = money(syn.get("awardCeiling"))
    floor = money(syn.get("awardFloor"))
    funding = money(syn.get("estimatedFunding"))
    awards = syn.get("numberOfAwards") or "n/a"
    theme_txt = ", ".join(themes) if themes else "general health"
    snippet = synopsis[:280].rstrip() + ("…" if len(synopsis) > 280 else "")
    text = (
        f"{number} ({agency}). {snippet} "
        f"Est. programme funding {funding}; award floor {floor}; ceiling {ceiling}; "
        f"~{awards} award(s). Entity types include: {types}. "
        f"FairBanks / FCHIP theme fit: {theme_txt}. "
        f"Ugandan private company path — confirm full NOFO eligibility and "
        f"foreign-entity rules on Simpler.Grants.gov before drafting. {simpler_url}"
    )
    # Keep description readable; URL is also in its own column.
    text = re.sub(r"\s+" + re.escape(simpler_url), "", text).strip()
    return truncate_words(text, 100)


def fit_note(themes: list[str], text: str) -> str:
    if "uganda" in text.lower():
        geo = "Uganda-named"
    elif re.search(r"partner countr|africa|global health", text, re.I):
        geo = "Africa / global partner"
    else:
        geo = "International"
    theme = themes[0] if themes else "health"
    return f"{geo} · {theme}"


def collect_candidates(sess: requests.Session) -> dict[str, dict[str, Any]]:
    """Return map of opportunity id -> search hit."""
    found: dict[str, dict[str, Any]] = {}
    for query in SEARCH_QUERIES:
        try:
            hits = search_grantsgov(sess, query, start=0, rows=25)
        except requests.RequestException as exc:
            print(f"Search failed for {query!r}: {exc}")
            continue
        for hit in hits:
            oid = str(hit.get("id") or "")
            if oid and oid not in found:
                found[oid] = hit
        # Optional Simpler API enrichment
        for item in search_simpler_api(sess, query):
            oid = str(item.get("opportunity_id") or item.get("legacy_opportunity_id") or "")
            if not oid:
                continue
            if oid not in found:
                found[oid] = {
                    "id": oid,
                    "number": item.get("opportunity_number"),
                    "title": item.get("opportunity_title"),
                    "agency": item.get("agency_name"),
                    "closeDate": item.get("close_date"),
                    "oppStatus": item.get("opportunity_status"),
                    "_from_simpler_api": True,
                    "_simpler_uuid": item.get("opportunity_id"),
                }
        time.sleep(0.35)
    return found


def evaluate_hit(
    sess: requests.Session,
    hit: dict[str, Any],
    url_cache: dict[str, str],
) -> dict[str, Any] | None:
    oid = hit.get("id")
    if hit.get("_from_simpler_api"):
        # Prefer Grants.gov numeric details when possible.
        return None

    detail = fetch_opportunity(sess, oid)
    if not detail:
        return None
    syn = detail.get("synopsis") or {}
    title = unescape(detail.get("opportunityTitle") or hit.get("title") or "")
    number = detail.get("opportunityNumber") or hit.get("number") or ""
    synopsis = unescape(syn.get("synopsisDesc") or "")
    elig_desc = unescape(syn.get("applicantEligibilityDesc") or "")
    agency_code = (syn.get("agencyCode") or hit.get("agencyCode") or "").strip()
    blob = " ".join(
        [
            title,
            number,
            synopsis,
            elig_desc,
            hit.get("agency") or "",
            syn.get("agencyName") or "",
            agency_code,
        ]
    )
    blob = re.sub(r"&nbsp;|&#\d+;|&\w+;", " ", blob)
    blob = re.sub(r"<[^>]+>", " ", blob)

    applicant_types = syn.get("applicantTypes") or []
    if not entity_ok(applicant_types):
        return None
    if not is_uganda_eligible(blob, applicant_types, agency_code=agency_code):
        return None
    themes = theme_hits(blob)
    if not themes:
        return None
    # Prefer field / community programmes over US biomedical discovery RFAs.
    if re.search(r"^(RFA-[A-Z]{2}-|PAR-\d|PA-\d)", number) and not (
        GEO_STRONG.search(blob) or GEO_AFRICA.search(blob) or agency_code in GLOBAL_HEALTH_AGENCIES
    ):
        return None

    url_cache[f"_id_{number}"] = str(oid)
    simpler_url = resolve_simpler_url(sess, number, url_cache)
    close_raw = hit.get("closeDate") or syn.get("responseDate") or ""
    # Prefer MM/DD/YYYY from search hit when present.
    deadline_sort = parse_us_date(hit.get("closeDate")) if hit.get("closeDate") else parse_us_date(syn.get("responseDate"))
    gender = GENDER_BASED if WOMEN_LED.search(blob) else MULTI_GENDER
    highlight = "00B0F0" if ("uganda" in blob.lower() or "surveillance" in themes[0].lower() or "digital" in themes[0].lower()) else None
    if "climate" in (themes[0] if themes else ""):
        highlight = "00B0F0"
    if gender == GENDER_BASED:
        highlight = highlight or "92D050"

    row = {
        "title": title,
        "url": simpler_url,
        "description": build_description(detail, themes, simpler_url),
        "deadline": format_deadline(hit.get("closeDate") or close_raw, syn.get("responseDateDesc")),
        "deadline_sort": deadline_sort,
        "gender": gender,
        "fit": fit_note(themes, blob),
        "opportunity_number": number,
        "agency": syn.get("agencyName") or hit.get("agency") or "",
        "award_floor": money(syn.get("awardFloor")),
        "award_ceiling": money(syn.get("awardCeiling")),
        "estimated_funding": money(syn.get("estimatedFunding")),
        "expected_awards": syn.get("numberOfAwards") or "Not listed",
        "applicant_types": "; ".join(
            t.get("description", "") for t in (syn.get("applicantTypes") or []) if str(t.get("id")) in ELIG_FOR_PROFIT
        ),
        "themes": "; ".join(themes),
        "grants_gov_id": str(oid),
        "highlight": highlight,
    }
    if word_count(row["description"]) > 100:
        row["description"] = truncate_words(row["description"], 100)
    return row


def write_workbook(rows: list[dict[str, Any]], scanned_queries: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    headers = [
        "Project Title",
        "URL",
        "Brief Description",
        "Application Deadline",
        "Gender Category",
        "Application Status",
        "Submission Status",
        "Submission Date",
        "Application Folder",
        "Opportunity Number",
        "Agency",
        "Est. Funding",
        "Award Floor",
        "Award Ceiling",
        "Expected Awards",
        "Eligible Entity Types (FB-relevant)",
        "Theme Tags",
    ]
    header_fill = PatternFill("solid", fgColor="0B3D2E")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    alt_fill = PatternFill("solid", fgColor="F3F8F5")
    gender_fill = PatternFill("solid", fgColor="F6E7F2")
    multi_fill = PatternFill("solid", fgColor="E7F0FA")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")

    rows_sorted = sorted(rows, key=lambda r: (r["deadline_sort"], r["title"]))

    for i, row in enumerate(rows_sorted, 2):
        highlight = row.get("highlight")
        values = [
            row["title"],
            row["url"],
            row["description"],
            row["deadline"],
            row["gender"],
            "Not started",
            "unknown",
            "",
            "",
            row.get("opportunity_number", ""),
            row.get("agency", ""),
            row.get("estimated_funding", ""),
            row.get("award_floor", ""),
            row.get("award_ceiling", ""),
            row.get("expected_awards", ""),
            row.get("applicant_types", ""),
            row.get("themes", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i, col, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if highlight:
                cell.fill = PatternFill("solid", fgColor=highlight)
            elif i % 2 == 0:
                cell.fill = alt_fill
        gcell = ws.cell(i, 5)
        if not highlight:
            gcell.fill = gender_fill if row["gender"] == GENDER_BASED else multi_fill
        gcell.font = Font(name="Calibri", size=10, bold=True)
        ws.row_dimensions[i].height = 110

    widths = {
        "A": 48,
        "B": 52,
        "C": 74,
        "D": 30,
        "E": 28,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 20,
        "J": 22,
        "K": 36,
        "L": 16,
        "M": 14,
        "N": 14,
        "O": 14,
        "P": 40,
        "Q": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 26
    last = len(rows_sorted) + 1
    ws.auto_filter.ref = f"A1:Q{last}"
    ws.freeze_panes = "A2"

    meta = wb.create_sheet("Scan Notes")
    gb = sum(1 for r in rows_sorted if r["gender"] == GENDER_BASED)
    mg = len(rows_sorted) - gb
    priority = [r["title"] for r in rows_sorted if r.get("highlight") in {"00B0F0", "92D050"}][:8]
    notes = [
        ("Scanned on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        (
            "Document type",
            "Live Simpler.Grants.gov / Grants.gov scan → new_grants.xlsx. "
            "Does not replace tracker.xlsx or new_opportunities.xlsx.",
        ),
        (
            "Applicant context",
            "FAIRBANKS MEDICAL CENTRE LIMITED — Ugandan private company "
            "limited by shares (TIN 1053370026); women-led (MD Racheal "
            "Nabukeera). Not an NGO/CBO. Live Community Reach + FCHIP MVP "
            "(CHW/VHT, GIS, climate, EMR/HMS APIs). Source: FairBanks "
            "Blueprint v1.0.1 in .cursor/source-of-truth/.",
        ),
        (
            "Sources",
            "https://simpler.grants.gov (listing URLs); public Grants.gov "
            "search2 + fetchOpportunity APIs for catalog data; optional "
            "SIMPLER_GRANTS_API_KEY for api.simpler.grants.gov.",
        ),
        (
            "Approach",
            "1) Search open Health (HL) opportunities. 2) Keep for-profit / "
            "small-business / unrestricted entity types. 3) Keep Uganda / "
            "Africa / global-partner geography. 4) Keep FairBanks theme fit "
            "(surveillance, community/maternal health, digital/AI, climate–"
            "health, HIV/infectious, health systems). 5) Resolve Simpler "
            "opportunity URLs. 6) Sort by earliest close date.",
        ),
        (
            "Search queries",
            "; ".join(scanned_queries),
        ),
        (
            "Eligibility filter",
            "Still posted + Health category + entity codes 22/23/99 + "
            "Uganda/Africa/global-partner text + FairBanks theme tags. "
            "US-domestic-only packages excluded. Always confirm foreign-"
            "entity and country rules on the official NOFO.",
        ),
        (
            "Gender categories",
            f"Gender-based (women/girls): {gb}. Multi-gender: {mg}. "
            f"Total: {len(rows_sorted)}.",
        ),
        (
            "Highlight legend",
            "Blue = strong Uganda / FCHIP / surveillance / digital fit. "
            "Green = women/girls framing. Unhighlighted = eligible but "
            "narrower theme match.",
        ),
        (
            "Priority soon",
            "; ".join(priority) if priority else "See rows sorted by deadline.",
        ),
        (
            "Important",
            "Federal listings change quickly. Confirm deadline, eligibility, "
            "and apply path on Simpler.Grants.gov before investing proposal "
            "time. FairBanks is a private company — NGO-only calls are out "
            "of scope unless a partner pathway is explicit.",
        ),
        ("Build script", "opportunities/build_new_grants.py"),
        ("Output file", "opportunities/new_grants.xlsx"),
        ("Related working tracker", "opportunities/tracker.xlsx"),
    ]
    meta["A1"] = "Field"
    meta["B1"] = "Detail"
    meta["A1"].font = Font(bold=True, color="FFFFFF")
    meta["B1"].font = Font(bold=True, color="FFFFFF")
    meta["A1"].fill = header_fill
    meta["B1"].fill = header_fill
    for i, (field, detail) in enumerate(notes, 2):
        meta.cell(i, 1, field).alignment = Alignment(vertical="top", wrap_text=True)
        meta.cell(i, 2, detail).alignment = Alignment(vertical="top", wrap_text=True)
        meta.row_dimensions[i].height = 70
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 96

    fit = wb.create_sheet("Fit Summary")
    fit_headers = [
        "Deadline sort",
        "Project Title",
        "FairBanks fit note",
        "Gender",
        "Opportunity Number",
        "Simpler URL",
    ]
    for col, h in enumerate(fit_headers, 1):
        cell = fit.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for i, row in enumerate(rows_sorted, 2):
        fit.cell(i, 1, row["deadline_sort"])
        fit.cell(i, 2, row["title"])
        fit.cell(i, 3, row.get("fit", ""))
        fit.cell(i, 4, row["gender"])
        fit.cell(i, 5, row.get("opportunity_number", ""))
        fit.cell(i, 6, row.get("url", ""))
        for c in range(1, 7):
            fit.cell(i, c).alignment = Alignment(wrap_text=True, vertical="top")
            fit.cell(i, c).border = thin
        fit.row_dimensions[i].height = 36
    fit.column_dimensions["A"].width = 14
    fit.column_dimensions["B"].width = 55
    fit.column_dimensions["C"].width = 36
    fit.column_dimensions["D"].width = 28
    fit.column_dimensions["E"].width = 22
    fit.column_dimensions["F"].width = 55

    wb.save(OUT)
    print(f"Wrote {OUT} with {len(rows_sorted)} grants")
    print(f"Gender-based: {gb}, Multi-gender: {mg}")


def main() -> None:
    sess = session()
    print("Collecting Simpler.Grants.gov / Grants.gov candidates…")
    candidates = collect_candidates(sess)
    print(f"Unique search hits: {len(candidates)}")

    url_cache: dict[str, str] = {}
    rows: list[dict[str, Any]] = []
    for idx, (oid, hit) in enumerate(candidates.items(), 1):
        try:
            row = evaluate_hit(sess, hit, url_cache)
        except requests.RequestException as exc:
            print(f"  skip {oid}: {exc}")
            row = None
        if row:
            rows.append(row)
            print(f"  [{idx}/{len(candidates)}] KEEP {row['opportunity_number']} — {row['deadline_sort']}")
        else:
            title = (hit.get("title") or hit.get("number") or oid)[:60]
            print(f"  [{idx}/{len(candidates)}] drop {title}")
        time.sleep(0.25)

    if not rows:
        raise SystemExit("No eligible grants found — check network / filters.")

    write_workbook(rows, SEARCH_QUERIES)


if __name__ == "__main__":
    main()

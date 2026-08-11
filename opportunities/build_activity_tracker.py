#!/usr/bin/env python3
"""Build opportunities/activity_tracker.xlsx from tracker + new_grants + today's actions."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "activity_tracker.xlsx"
TRACKER = ROOT / "tracker.xlsx"
NEW_GRANTS = ROOT / "new_grants.xlsx"

TODAY = "11 August 2026"
HEADER_FILL = PatternFill("solid", fgColor="0B3D2E")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
ALT = PatternFill("solid", fgColor="F3F8F5")
URGENT = PatternFill("solid", fgColor="FCE4D6")
TODAY_FILL = PatternFill("solid", fgColor="FFF2CC")
DONE = PatternFill("solid", fgColor="C6EFCE")
BLUE = PatternFill("solid", fgColor="DDEBF7")


def style_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_row(ws, row: int, values: list, fill=None) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row, col, val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN
        cell.font = Font(name="Calibri", size=10)
        if fill:
            cell.fill = fill
        elif row % 2 == 0:
            cell.fill = ALT


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_letter = col
        ws.column_dimensions[col].width = width


def read_tracker_rows() -> list[dict]:
    wb = load_workbook(TRACKER, data_only=True)
    ws = wb["Opportunities"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        rows.append(
            {
                "title": r[0],
                "url": r[1] or "",
                "deadline": r[3] or "",
                "app_status": r[5] or "",
                "sub_status": r[6] or "",
                "folder": r[8] or "",
                "source": "tracker.xlsx",
            }
        )
    return rows


def read_grant_rows() -> list[dict]:
    wb = load_workbook(NEW_GRANTS, data_only=True)
    ws = wb["Opportunities"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        rows.append(
            {
                "title": r[0],
                "url": r[1] or "",
                "deadline": r[3] or "",
                "app_status": r[5] or "Not started",
                "opp_number": r[9] if len(r) > 9 else "",
                "agency": r[10] if len(r) > 10 else "",
                "themes": r[16] if len(r) > 16 else "",
                "source": "new_grants.xlsx",
            }
        )
    return rows


def build() -> None:
    tracker = read_tracker_rows()
    grants = read_grant_rows()

    wb = Workbook()

    # --- Sheet 1: Today's actions ---
    today = wb.active
    today.title = "Today"
    today_headers = [
        "When",
        "Owner",
        "Activity",
        "Related opportunity / asset",
        "Link / path",
        "Status",
        "Notes / next step",
    ]
    style_header(today, today_headers)

    today_rows = [
        [
            "08:00",
            "Racheal + team",
            "Meeting: collectively prepare and apply for CDC-RFA-JG-26-0056 (detect / notify / respond globally)",
            "CDC GHS Global 0056",
            "https://simpler.grants.gov/opportunity/8454e463-cd43-4d0d-97a2-8a4310e0ce6b",
            "Scheduled",
            "Use applications/cdc-ghs-global pack + application_answers.md. Confirm SAM/UEI, local-partner letter, budgets.",
        ],
        [
            "Today",
            "Racheal",
            "Contact UVRI for possible collaboration (surveillance / research / outbreak partnership)",
            "CDC GHS Global 0056 + FairBanks Community Reach",
            "applications/cdc-ghs-global",
            "To do",
            "Ask for a named contact and LOI pathway that can strengthen the 0056 application.",
        ],
        [
            "Today",
            "Racheal",
            "Contact IDI (Infectious Diseases Institute) for possible collaboration",
            "CDC GHS Global 0056 + FairBanks Community Reach",
            "applications/cdc-ghs-global",
            "To do",
            "Explore clinical / surveillance / training partnership; request intro call this week.",
        ],
        [
            "Today",
            "Team (collective)",
            "Go through FairBanks Management Toolkit v1.0.0 together",
            "Institutional forms companion to Handbook v1.0.1",
            ".cursor/source-of-truth/ (Management Toolkit v1.0.0)",
            "To do",
            "Agree which forms to use for 0056 evidence, partner notes, and daily ops.",
        ],
        [
            "Today",
            "Wilson / ops",
            "Refuel Cursor subscription so FCHIP MVP development can continue",
            "FCHIP MVP",
            "Cursor billing / subscription",
            "To do",
            "Confirm plan active and developer access restored after payment.",
        ],
    ]
    for i, row in enumerate(today_rows, 2):
        fill = TODAY_FILL if row[5] == "Scheduled" else URGENT
        style_row(today, i, row, fill=fill)
        today.row_dimensions[i].height = 72
    set_widths(
        today,
        {"A": 12, "B": 18, "C": 55, "D": 36, "E": 52, "F": 12, "G": 48},
    )
    today.freeze_panes = "A2"
    today.auto_filter.ref = f"A1:G{len(today_rows) + 1}"
    today.row_dimensions[1].height = 28

    # --- Sheet 2: Activity log (mix of grant follow-ups + ops) ---
    log = wb.create_sheet("Activity log")
    log_headers = [
        "Priority",
        "Due / when",
        "Owner",
        "Activity",
        "Type",
        "Related title",
        "Deadline",
        "App status",
        "Folder / URL",
        "Status",
        "Source",
    ]
    style_header(log, log_headers)

    activities: list[list] = []

    # Today first
    for row in today_rows:
        activities.append(
            [
                "P0 Today",
                f"{TODAY} {row[0]}" if row[0] != "Today" else TODAY,
                row[1],
                row[2],
                "Ops / partnership / apply",
                row[3],
                "14 August 2026" if "0056" in row[2] or "0056" in row[3] else "",
                "Drafting" if "0056" in (row[2] + row[3]) else "",
                row[4],
                row[5],
                "manual today",
            ]
        )

    # Open / drafting tracker items that still need work
    for t in tracker:
        status = (t["app_status"] or "").lower()
        if status in {"complete", "submitted poc"} and "0056" not in (t["title"] or ""):
            continue
        if "0056" in (t["title"] or ""):
            continue  # already covered in today block as apply meeting
        priority = "P1 Active"
        if "drafting" in status or "mvp" in status:
            priority = "P1 Active"
        activities.append(
            [
                priority,
                "This week",
                "Team",
                f"Advance application / pack: {t['title'][:80]}",
                "Grant / opportunity",
                t["title"],
                t["deadline"],
                t["app_status"],
                t["folder"] or t["url"],
                "Open",
                t["source"],
            ]
        )

    # new_grants follow-ups (skip 0054 if complete on tracker; highlight open ones)
    tracker_titles = " ".join((t["title"] or "").lower() for t in tracker)
    for g in grants:
        title = g["title"] or ""
        # Skip Uganda 0054 if already Complete on working tracker
        if "uganda" in title.lower() and "local partnerships" in title.lower() and "0054" in tracker_titles:
            if "complete" in tracker_titles:
                pass
        # Always list as grant-pipeline follow-up if not complete on tracker
        already = any(
            (g.get("opp_number") or "") and (g.get("opp_number") or "") in (t["title"] or "")
            for t in tracker
        )
        # Map known overlaps
        on_tracker = False
        for t in tracker:
            if g.get("opp_number") and g["opp_number"] in (t["title"] or ""):
                on_tracker = True
                if (t["app_status"] or "").lower() == "complete":
                    activities.append(
                        [
                            "P3 Watch",
                            g["deadline"],
                            "Team",
                            f"Monitor / archive: {title[:70]}",
                            "Grant / opportunity",
                            title,
                            g["deadline"],
                            t["app_status"],
                            t.get("folder") or g["url"],
                            "Complete on tracker",
                            "new_grants.xlsx + tracker.xlsx",
                        ]
                    )
                elif "0056" in (g.get("opp_number") or ""):
                    # Covered by today apply meeting
                    continue
                else:
                    activities.append(
                        [
                            "P1 Active",
                            g["deadline"],
                            "Team",
                            f"Continue drafting: {title[:70]}",
                            "Grant / opportunity",
                            title,
                            g["deadline"],
                            t["app_status"],
                            t.get("folder") or g["url"],
                            "Open",
                            "new_grants.xlsx + tracker.xlsx",
                        ]
                    )
                break
        else:
            # Not on tracker yet
            due = g["deadline"] or ""
            priority = "P2 Pipeline"
            if "14 August" in due or "August 2026" in due:
                priority = "P1 Active"
            activities.append(
                [
                    priority,
                    due or "TBD",
                    "Team",
                    f"Triage / decide whether to apply: {title[:70]}",
                    "Grant / opportunity",
                    title,
                    due,
                    g["app_status"],
                    g["url"],
                    "Not started",
                    "new_grants.xlsx",
                ]
            )

    # Sort: P0 first, then P1, then by due text
    order = {"P0 Today": 0, "P1 Active": 1, "P2 Pipeline": 2, "P3 Watch": 3}
    activities.sort(key=lambda a: (order.get(a[0], 9), a[1], a[5]))

    for i, row in enumerate(activities, 2):
        fill = None
        if row[0] == "P0 Today":
            fill = URGENT if row[9] != "Scheduled" else TODAY_FILL
        elif row[0] == "P1 Active":
            fill = BLUE
        elif row[9] == "Complete on tracker":
            fill = DONE
        style_row(log, i, row, fill=fill)
        log.row_dimensions[i].height = 58

    set_widths(
        log,
        {
            "A": 12,
            "B": 22,
            "C": 14,
            "D": 52,
            "E": 18,
            "F": 40,
            "G": 28,
            "H": 14,
            "I": 42,
            "J": 16,
            "K": 22,
        },
    )
    log.freeze_panes = "A2"
    log.auto_filter.ref = f"A1:K{len(activities) + 1}"
    log.row_dimensions[1].height = 28

    # --- Sheet 3: Grant pipeline snapshot ---
    pipe = wb.create_sheet("Grant pipeline")
    pipe_headers = [
        "Source",
        "Title",
        "Opportunity #",
        "Deadline",
        "App status",
        "Folder",
        "URL",
        "Action needed",
    ]
    style_header(pipe, pipe_headers)
    pipe_rows: list[list] = []

    for t in tracker:
        opp = ""
        for token in (t["title"] or "").replace("(", " ").replace(")", " ").split():
            if token.startswith("CDC-RFA") or token.startswith("DFOP") or token.startswith("W81"):
                opp = token
                break
        action = "Maintain / archive"
        st = (t["app_status"] or "").lower()
        if "draft" in st:
            action = "Finish pack and submit"
        elif "mvp" in st:
            action = "Complete MVP / pitch follow-up"
        elif st == "complete":
            action = "No action unless funder replies"
        pipe_rows.append(
            [
                "tracker.xlsx",
                t["title"],
                opp,
                t["deadline"],
                t["app_status"],
                t["folder"],
                t["url"],
                action,
            ]
        )

    for g in grants:
        # Avoid duplicate if already on tracker by opp number in title
        dup = False
        for t in tracker:
            if g.get("opp_number") and g["opp_number"] in (t["title"] or ""):
                dup = True
                break
        if dup:
            continue
        pipe_rows.append(
            [
                "new_grants.xlsx",
                g["title"],
                g.get("opp_number") or "",
                g["deadline"],
                g["app_status"],
                "",
                g["url"],
                "Triage fit; promote to tracker if pursuing",
            ]
        )

    # Soonest deadline-ish: keep Aug 2026 near top by simple string sort on deadline
    pipe_rows.sort(key=lambda r: (r[3] or "9999", r[1] or ""))

    for i, row in enumerate(pipe_rows, 2):
        fill = URGENT if "14 August 2026" in (row[3] or "") and "draft" in (row[4] or "").lower() else None
        if fill is None and "14 August" in (row[3] or "") and row[0] == "new_grants.xlsx":
            fill = BLUE
        style_row(pipe, i, row, fill=fill)
        pipe.row_dimensions[i].height = 56

    set_widths(
        pipe,
        {"A": 16, "B": 52, "C": 20, "D": 30, "E": 16, "F": 28, "G": 48, "H": 36},
    )
    pipe.freeze_panes = "A2"
    pipe.auto_filter.ref = f"A1:H{len(pipe_rows) + 1}"

    # --- Sheet 4: Notes ---
    notes = wb.create_sheet("Notes")
    style_header(notes, ["Field", "Detail"])
    note_rows = [
        ("Built on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Activity date focus", TODAY),
        (
            "Purpose",
            "Day-to-day activity tracker combining opportunities/tracker.xlsx, "
            "opportunities/new_grants.xlsx, and live ops (partners, toolkit, tools).",
        ),
        (
            "Today priority",
            "08:00 meeting to apply for CDC-RFA-JG-26-0056; Racheal contacts UVRI and IDI; "
            "team reviews FairBanks Management Toolkit v1.0.0; refill Cursor subscription for MVP.",
        ),
        (
            "0056 listing",
            "https://simpler.grants.gov/opportunity/8454e463-cd43-4d0d-97a2-8a4310e0ce6b",
        ),
        (
            "0056 pack",
            "applications/cdc-ghs-global (documents + application_answers.md)",
        ),
        (
            "Toolkit",
            ".cursor/source-of-truth/ FairBanks Management Toolkit v1.0.0 "
            "(companion to Handbook v1.0.1)",
        ),
        (
            "Partners to contact",
            "UVRI (Uganda Virus Research Institute); IDI (Infectious Diseases Institute).",
        ),
        (
            "How to update",
            "Edit Today and Activity log sheets as work finishes. Re-run this builder "
            "only when you want a fresh merge from tracker/new_grants; keep manual "
            "status notes if you edit by hand.",
        ),
        ("Builder", "opportunities/build_activity_tracker.py"),
        ("Related files", "opportunities/tracker.xlsx · opportunities/new_grants.xlsx"),
    ]
    for i, (field, detail) in enumerate(note_rows, 2):
        style_row(notes, i, [field, detail])
        notes.row_dimensions[i].height = 48
    set_widths(notes, {"A": 22, "B": 96})

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Today rows: {len(today_rows)}")
    print(f"Activity log rows: {len(activities)}")
    print(f"Pipeline rows: {len(pipe_rows)}")


if __name__ == "__main__":
    build()

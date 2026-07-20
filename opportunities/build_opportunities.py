#!/usr/bin/env python3
"""Build the canonical FairBanks opportunity tracker Excel.

Writes opportunities/opportunities.xlsx per
opportunities/rules/source_of_truth.mdc.

Includes the former peri-urban / FCHIP complementary scan (merged July 2026).
Only Uganda-eligible, still-open opportunities; gender-tagged.
Row highlights (green/blue/red/black) mark peri-urban priority fits.
"""

from datetime import datetime
from pathlib import Path
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent
DEFAULT_OUT = ROOT / "opportunities.xlsx"

GENDER_BASED = "Gender-based (women/girls)"
MULTI_GENDER = "Multi-gender (all genders)"

# app_slug maps to applications/{slug}/ win-win packs.
OPPORTUNITIES = [
    {
        "title": "AWIEF Pitch n Grow 2026 — Women-Led Deep-Tech Startup Competition",
        "url": "https://opportunitiesforyouth.org/2026/06/25/awief-pitch-n-grow-2026-applications-open-for-africas-leading-women-led-deep-technology-startup-competition/",
        "app_slug": "awief",
        "gender": GENDER_BASED,
        "description": (
            "Continent-wide competition for women-led ventures using deep "
            "technology under the theme 'Deep Roots. Digital Futures.' Tracks: "
            "Idea, Startup, Scale-Up. HealthTech (digital healthcare, "
            "telemedicine, AI diagnostics) and AI/ML are priority sectors, so "
            "FCHIP is a strong fit. Benefits include a training bootcamp, "
            "mentorship, pitch coaching, investor visibility, media exposure, "
            "travel support for finalists, and cash prizes. Open to women "
            "founders across Africa, so Ugandans qualify. Submit before "
            "screening begins."
        ),
        "deadline": "20 July 2026",
        "deadline_sort": "2026-07-20",
    },
    {
        "title": "The Gadfly Project Custom Web Application Grant 2026",
        "url": "https://opportunitiesforyouth.org/2026/07/16/the-gadfly-project-custom-web-application-grant-2026-apply-for-in-kind-software-development-support-worth-up-to-us100000/",
        "app_slug": "gadfly",
        "gender": MULTI_GENDER,
        "description": (
            "In-kind grant giving mission-driven organisations professional "
            "software development (valued US$5,000–100,000) instead of cash. "
            "Gadfly's in-house team helps you build custom web apps, databases, "
            "outreach platforms, or internal systems that expand community "
            "impact. Open to nonprofits, charities, social enterprises, and "
            "community organisations worldwide, so a Ugandan health "
            "organisation qualifies. Reviewed quarterly; selection weighs "
            "users served and community impact over technical complexity. "
            "Useful for building FCHIP mobile and data tools."
        ),
        "deadline": "24 July 2026 (next cycle 23 October 2026)",
        "deadline_sort": "2026-07-24",
    },
    {
        "title": "AUC Venture Lab Women Innovation Fellowship 2026",
        "url": "https://opportunitiesforyouth.org/2026/07/16/auc-venture-lab-women-innovation-fellowship-empowering-women-entrepreneurs-in-egypt-and-beyond/",
        "app_slug": "auc",
        "gender": GENDER_BASED,
        "description": (
            "American University in Cairo fellowship (1 Sep–31 Oct 2026, "
            "Cairo) empowering women entrepreneurs and innovators through "
            "leadership training, ecosystem access, mentorship, and "
            "cross-regional exposure across Egypt and Denmark. Open to women "
            "across sectors including HealthTech, AI, and sustainability, and "
            "to founders beyond Egypt, so Ugandan women innovators may apply. "
            "Strong networking and investment-readiness support for a "
            "woman-led health venture."
        ),
        "deadline": "25 July 2026",
        "deadline_sort": "2026-07-25",
    },
    {
        "title": "Young Feminist AI School 2026 (UN Women)",
        "url": "https://opportunitiesforyouth.org/2026/07/16/ai-for-gender-equality-un-women-ai-school-opens-for-changemakers/",
        "app_slug": "feminist-ai",
        "gender": GENDER_BASED,
        "description": (
            "Free, fully online 12-week UN Women programme (Aug–Oct 2026) "
            "teaching practical, no-code AI skills for advocacy, policy, civic "
            "tech, and responsible AI. Open to young people aged 18–30 "
            "worldwide, with at least 70% of places reserved for young women "
            "and girls. About four hours weekly; every participant builds a "
            "real project. Good capacity building for team members applying "
            "gender-responsive AI to community health. Ugandans eligible."
        ),
        "deadline": "28 July 2026",
        "deadline_sort": "2026-07-28",
    },
    {
        "title": "GirlCode Hackathon 2026 (includes Kampala, Uganda)",
        "url": "https://opportunitiesforyouth.org/2026/07/15/building-africas-next-generation-of-ai-powered-female-innovators-apply-for-the-2026-girlcode-hackathon-across-six-african-countries/",
        "app_slug": "girlcode",
        "gender": GENDER_BASED,
        "description": (
            "Multi-country AI hackathon for female innovators across South "
            "Africa, Kenya, Uganda, Tanzania, and Ghana, building Africa's "
            "next generation of women in tech. The Kampala event runs 5–6 "
            "September 2026 at UICT Nakawa. Great for visibility, "
            "team-building, recruiting women technologists, or showcasing "
            "FCHIP concepts. No single formal deadline is stated, so register "
            "before the local event date. Ugandan women eligible."
        ),
        "deadline": "Not formally stated — register before Kampala event 5–6 Sep 2026",
        "deadline_sort": "2026-09-05",
    },
    {
        "title": "U.S. Dept of State — Uganda Health System MOU (up to $60M)",
        "url": "https://opportunitiesforyouth.org/2026/06/15/u-s-department-of-state-launches-up-to-60-million-funding-opportunity-to-strengthen-ugandas-health-system-through-the-health-foreign-assistance-mou-implementation-plan/",
        "app_slug": "dos-uganda",
        "gender": MULTI_GENDER,
        "description": (
            "Major U.S. Department of State call (DFOP0017890) to strengthen "
            "Uganda's health system under the U.S.–Uganda Health MOU: up to "
            "$60M across ~15 awards over five years. Objective 4 funds scaling "
            "digital health systems (EMRs, DHIS2, eCHIS, iHRIS, national data "
            "warehouse); other tracks cover community-based health financing, "
            "faith/community facilities, and supply-chain accountability. "
            "Two-step process starting with a Statement of Interest. Highly "
            "aligned with FCHIP and FairBanks community health work. Ugandan "
            "organisations are the intended applicants."
        ),
        "deadline": "31 July 2026 (Statement of Interest, 5:00 PM EDT)",
        "deadline_sort": "2026-07-31",
    },
    {
        "title": "World Health Summit 2026 Youth Group Opportunity (Berlin)",
        "url": "https://opportunitiesforyouth.org/2026/06/26/apply-now-world-health-summit-2025-stipend-program-for-global-health-changemakers/",
        "app_slug": "whs",
        "gender": MULTI_GENDER,
        "description": (
            "Complimentary group tickets (3–15 per organisation) for "
            "youth-led organisations to attend all three days of the World "
            "Health Summit 2026 in Berlin. Open to youth-led groups working "
            "in health, advocacy, education, or social impact, with nominees "
            "aged 28 or under; travel, visa, and accommodation are "
            "self-funded. Diverse and underrepresented regions are "
            "encouraged, so a Ugandan community-health organisation is a "
            "strong candidate for global visibility and networking."
        ),
        "deadline": "31 July 2026",
        "deadline_sort": "2026-07-31",
    },
    {
        "title": "African Union Youth SRH Ambassadors Initiative 2026 (AUDA-NEPAD)",
        "url": "https://opportunitiesforyouth.org/2026/06/13/african-union-youth-srh-ambassadors-initiative-2026-apply-to-become-an-auda-nepad-youth-srh-ambassador/",
        "app_slug": "auda-srh",
        "gender": MULTI_GENDER,
        "description": (
            "African Union (AUDA-NEPAD) initiative recruiting young Africans "
            "aged 18–35 as voluntary Sexual and Reproductive Health "
            "ambassadors for advocacy, awareness, and community mobilisation. "
            "Offers recognition, training, mentorship, networking, and "
            "sponsored participation in selected events (no salary). "
            "Applicants should show SRH commitment and an active digital "
            "presence; diverse genders and regions are encouraged. Relevant "
            "to FairBanks community reach and youth health engagement. "
            "Ugandans eligible — note the deadline is today."
        ),
        "deadline": "17 July 2026 (closing)",
        "deadline_sort": "2026-07-17",
    },
    {
        "title": "Brandtech Group AI for Good Scholarship 2026 (One Young World)",
        "url": "https://opportunitiesforyouth.org/2026/07/15/ai-for-good-scholarship-2025-empowering-young-innovators-to-address-global-challenges/",
        "app_slug": "oyw",
        "gender": MULTI_GENDER,
        "description": (
            "Fully funded scholarship for five young African leaders (aged "
            "18–30) using AI for social, health, education, environmental, or "
            "civic impact, to attend the One Young World Summit 2026 in Cape "
            "Town. Healthcare is a named priority area. Requires African "
            "nationality and residency plus a proven record of responsible AI "
            "work. A strong platform for a founder or team member advancing AI "
            "in community health. Ugandans eligible."
        ),
        "deadline": "2 August 2026",
        "deadline_sort": "2026-08-02",
    },
    {
        "title": "Africa CDC African Epidemic Services (AES) Epidemiology Fellowship 2026–2028",
        "url": "https://opportunitiesforyouth.org/2026/07/16/africa-cdc-african-epidemic-services-aes-epidemiology-fellowship-2026-2028-fully-funded-apply-by-26-august-2026/",
        "app_slug": "africa-cdc",
        "gender": MULTI_GENDER,
        "description": (
            "Fully funded two-year Africa CDC fellowship — three months' "
            "training in Addis Ababa plus 21 months' field placement — for "
            "citizens of African Union member states under 35 already employed "
            "in public health, with a relevant degree and three years' "
            "experience. Builds outbreak investigation, surveillance, and "
            "data-analysis skills; a Public Health Informatics track also "
            "exists. Covers stipend, travel, insurance, hardware, and "
            "software. Excellent professional development for FairBanks "
            "epidemiology or informatics talent. Ugandans eligible."
        ),
        "deadline": "26 August 2026",
        "deadline_sort": "2026-08-26",
    },
    {
        "title": "IFAD Grant Call — Rural Sector Performance Assessment (up to US$1.2M)",
        "url": "https://opportunitiesforyouth.org/2026/07/14/ifad-announces-us1-2-million-grant-call-for-proposals-to-strengthen-rural-development-policies-worldwide/",
        "app_slug": "ifad",
        "gender": MULTI_GENDER,
        "description": (
            "IFAD call for proposals worth up to US$1.2M to implement "
            "'Leveraging IFAD's Rural Sector Performance Assessment for Policy "
            "and Investment.' Open to non-profits, private companies, "
            "universities, think tanks, and research institutes (20–25% "
            "counterpart contribution). Optional non-binding Expression of "
            "Interest by 31 July 2026. Best fit if FairBanks partners on rural "
            "policy and data with livelihood–health linkages; verify thematic "
            "match before investing proposal effort. Ugandan organisations "
            "eligible."
        ),
        "deadline": "4 September 2026 (12:00 CET); optional EOI 31 July 2026",
        "deadline_sort": "2026-09-04",
    },
    {
        "title": "World Bank GovTech Innovation Challenge 2026",
        "url": "https://opportunitiesforyouth.org/2026/07/10/world-bank-govtech-innovation-challenge-2026-global-call-for-proposals-to-build-digital-solutions-for-governments/",
        "app_slug": "govtech",
        "gender": MULTI_GENDER,
        "description": (
            "Global World Bank / SECO / Trust Valley challenge connecting "
            "technology companies with government digitalisation needs. "
            "Selected innovators get fully funded capacity building, a "
            "Switzerland bootcamp, and a chance to build a proof of concept "
            "with a government partner. Current open challenges are Moroccan "
            "public-finance audit use cases, but new calls appear over time — "
            "watch for health or Africa challenges matching district health "
            "intelligence. Deadlines vary by challenge. Companies worldwide, "
            "including Ugandan, may apply."
        ),
        "deadline": "Varies by challenge — check the specific challenge page",
        "deadline_sort": "9999-12-31",
    },
    # --- Former peri-urban / FCHIP complementary scan (merged Jul 2026) ---
    {
        "title": "UN Women — Spotlight Initiative Africa Regional Programme (SIARP) 2.0",
        "url": "https://africa.unwomen.org/en/programme-implementation/2026/06/call-for-proposals-for-civil-society-partnerships-for-implementation-of-the-spotlight-initiative-africa-regional-programme-siarp-20",
        "gender": GENDER_BASED,
        "highlight": "92D050",
        "description": (
            "UN Women call for African civil society partners under SIARP 2.0. "
            "Grants of about US$110,000–548,000 for up to 27 months (from Oct "
            "2026) on ending violence against women and girls, harmful "
            "practices, SRHR, and women- and youth-led movements. Strongest "
            "fit if FairBanks joins a regional consortium linking peri-urban "
            "maternal/adolescent health, CHW outreach, and community "
            "accountability. Open to CSOs, women’s rights and youth-led "
            "groups across Africa — Ugandan organisations may apply; confirm "
            "regional/multi-country scope on the call page."
        ),
        "deadline": "27 July 2026 (5:00 PM East Africa Time)",
        "deadline_sort": "2026-07-27",
    },
    {
        "title": "AFNet Flexible Grant (African Women for Change Network)",
        "url": "https://afwcnet.org/women-grants",
        "gender": GENDER_BASED,
        "highlight": "92D050",
        "description": (
            "Small flexible grant (about US$5,000) for women leaders and "
            "women-led organisations driving community change in Africa. "
            "Reviewed on a rolling basis. Useful seed support for a "
            "woman-led FairBanks Community Reach or FCHIP pilot activity in "
            "peri-urban Kampala communities (outreach, maternal support, or "
            "CHW coordination). Confirm current application steps and "
            "eligibility on the AFNet channel before applying; listing "
            "details can change."
        ),
        "deadline": "Rolling",
        "deadline_sort": "9999-12-31",
    },
    {
        "title": "Africa Health-Tech Accelerator 2026 (Africa Health Tech / Africa CDC / Health ExCon)",
        "url": "https://accelerator.africahealthexcon.com/apply",
        "gender": MULTI_GENDER,
        "highlight": "00B0F0",
        "description": (
            "Six-month pan-African accelerator for early-stage health-tech "
            "startups and SMEs. Tracks include digital healthcare, health "
            "data and intelligent technologies, prevention/access services, "
            "and related innovation. Benefits: structured training, "
            "mentorship, seed-funding opportunities, investor Demo Day, and "
            "continental visibility. Needs at least two founders, an MVP, "
            "and a scale-across-Africa commitment. Excellent near-term fit "
            "for FCHIP as a community health intelligence / digital health "
            "venture. Deadline extended to 20 July 2026 — apply immediately."
        ),
        "deadline": "20 July 2026",
        "deadline_sort": "2026-07-20",
    },
    {
        "title": "Nexa Climate and Health Initiative (Grand Challenges Canada / Science for Africa Foundation)",
        "url": "https://www.grandchallenges.ca/rfp-nexa/",
        "app_slug": "nexa-climate-health",
        "gender": MULTI_GENDER,
        "highlight": "00B0F0",
        "description": (
            "Major open competition for climate–health innovations in "
            "Africa (and LAC for scale). Funds tools that turn climate "
            "risk signals into timely health action on mosquito-borne "
            "disease (e.g. malaria/dengue), extreme heat, and air "
            "quality. Proof of Concept up to US$200,000; Transition to "
            "Scale about US$250,000–2,000,000. PoC applicants must be "
            "incorporated in an eligible African country (Uganda "
            "included). Closest deep-tech match for FCHIP disease early "
            "warning in peri-urban communities. Submit via Fluxx by "
            "22 July 2026, 2:00 p.m. ET."
        ),
        "deadline": "22 July 2026 (2:00 p.m. ET / 6:00 p.m. UTC)",
        "deadline_sort": "2026-07-22",
    },
    {
        "title": "IDRC / GAC / CIHR — ANeSA Cohort 2 Letters of Interest (SRHR in sub-Saharan Africa)",
        "url": "https://idrc-crdi.ca/en/funding/call-letters-interest-addressing-neglected-areas-sexual-and-reproductive-health-and-0",
        "gender": MULTI_GENDER,
        "description": (
            "Open call for Letters of Interest under Addressing Neglected "
            "Areas of SRHR in sub-Saharan Africa (ANeSA). Up to six grants "
            "of up to CAD 1.2 million each (to 36 months) for "
            "gender-transformative implementation research with "
            "under-served populations. Priorities include family planning, "
            "adolescent SRHR, SGBV services, and SRHR advocacy. Led by "
            "SSA-based researchers; Uganda is an eligible country "
            "(additional IDRC approval may apply). Strong research "
            "partnership path for FairBanks peri-urban maternal and "
            "adolescent community health work."
        ),
        "deadline": "23 August 2026 (23:59 ET)",
        "deadline_sort": "2026-08-23",
    },
    {
        "title": "Gilead Global Public Health Awards 2026",
        "url": "https://opportunitiesforyouth.org/2026/05/25/gilead-global-public-health-awards-2026-100000-research-funding-for-early-career-investigators-in-hiv-and-viral-hepatitis/",
        "gender": MULTI_GENDER,
        "highlight": "FF0000",
        "description": (
            "Four awards of up to US$100,000 over two years for early-career "
            "investigators in LMICs (including Africa) focused on HIV and "
            "viral hepatitis public health / implementation science. "
            "Applicants need a terminal degree within 10 years, an "
            "in-country mentor, and about 50% research time. Best if a "
            "FairBanks-linked clinician or researcher studies community "
            "prevention, access, or peri-urban care pathways. Confirm on "
            "the official Gilead awards portal before applying."
        ),
        "deadline": "28 August 2026 (11:59 PM CEST)",
        "deadline_sort": "2026-08-28",
    },
    {
        "title": "Social Shifters Global Innovation Challenge 2026",
        "url": "https://www.socialshifters.co/global-innovation-challenge/",
        "gender": MULTI_GENDER,
        "highlight": "00B0F0",
        "description": (
            "Global challenge for youth-led projects and startups aligned "
            "with at least one SDG, offering up to about US$15,000 in "
            "support. Suitable for a young FairBanks / FCHIP team member "
            "piloting a community health, digital CHW, or peri-urban "
            "prevention idea. Lower funding than major grants but useful "
            "for visibility and early traction. Deadline is 31 August 2026, "
            "5pm UTC. Confirm age rules and award tracks on the Social "
            "Shifters challenge page before investing proposal time."
        ),
        "deadline": "31 August 2026 (5:00 PM UTC)",
        "deadline_sort": "2026-08-31",
    },
    {
        "title": "Fund for Innovation in Development (FID) — France (platform reopens 1 Sep 2026)",
        "url": "https://fundinnovation.dev/en/launch-project",
        "gender": MULTI_GENDER,
        "highlight": "000000",
        "description": (
            "French open innovation fund for solutions reducing poverty "
            "and inequality in ODA-eligible countries (Uganda eligible). "
            "Health is a priority sector alongside education, climate, and "
            "gender. Staged grants from pilot to scale (historically up to "
            "about €4 million at scale). Portal is closed for a technical "
            "upgrade and reopens 1 September 2026 for most stages "
            "(Preparation grants later). Strong longer-horizon fit for "
            "evidence-backed FCHIP pilots in peri-urban communities."
        ),
        "deadline": "Reopens 1 September 2026 — then rolling",
        "deadline_sort": "2026-09-01",
    },
    {
        "title": (
            "Global Health EDCTP3 — Enhancing integrated research and "
            "healthcare in SSA through digital innovation and AI "
            "(HORIZON-JU-GH-EDCTP3-2026-03-DIGIT-02)"
        ),
        "url": (
            "https://ec.europa.eu/info/funding-tenders/opportunities/portal/"
            "screen/opportunities/topic-details/"
            "HORIZON-JU-GH-EDCTP3-2026-03-DIGIT-02"
        ),
        "gender": MULTI_GENDER,
        "highlight": "00B0F0",
        "description": (
            "EU Global Health EDCTP3 Coordination and Support Action "
            "(about €18M topic budget; up to about €2.25M per project) to "
            "scale and coordinate existing digital health and AI tools "
            "across sub-Saharan Africa — interoperability, capacity, "
            "epidemic preparedness, and decision support for health "
            "workers. Not for building brand-new tech alone; best as an "
            "Africa–Europe consortium partner bringing FCHIP / peri-urban "
            "CHW data use-cases. Uganda/SSA organisations can join "
            "eligible consortia. Deadline 2 September 2026, 17:00 Brussels."
        ),
        "deadline": "2 September 2026 (17:00 Brussels time)",
        "deadline_sort": "2026-09-02",
    },
    {
        "title": "Wellcome Snakebite Innovation Prize (Challenge Works)",
        "url": "https://snakebiteprize.challenges.org/",
        "gender": MULTI_GENDER,
        "highlight": "FF0000",
        "description": (
            "£6.25 million challenge prize for solutions that strengthen "
            "community responses, speed access to care, or improve "
            "treatment delivery for snakebite in high-burden settings "
            "(including sub-Saharan Africa). Launch track ~£75,000 and "
            "Growth track ~£100,000 for finalists; larger awards later. "
            "Open worldwide until 16 September 2026, 12:00 UTC. Indirect "
            "fit if FairBanks adapts CHW alert/referral tools for "
            "emergency community response in underserved areas — less "
            "central to peri-urban Kampala NCDs/maternal work."
        ),
        "deadline": "16 September 2026 (12:00 UTC)",
        "deadline_sort": "2026-09-16",
    },
    {
        "title": "DIV Fund (Development Innovation Ventures) — Open Innovation Grants",
        "url": "https://www.div.fund/apply",
        "gender": MULTI_GENDER,
        "highlight": "00B0F0",
        "description": (
            "Evidence-driven open innovation fund accepting applications "
            "year-round. Stages roughly: Pilot up to ~US$200,000; Testing "
            "higher; Scale up to about US$1.5 million. Funds innovations "
            "that can cost-effectively improve millions of lives in "
            "low- and middle-income countries. Open to nonprofits, "
            "startups, universities, and social enterprises — Ugandan "
            "applicants eligible. Good medium-term pathway for FCHIP if "
            "you can show a clear pilot plan, cost-effectiveness case, "
            "and path to scale across peri-urban and district health "
            "systems."
        ),
        "deadline": "Rolling (applications accepted year-round)",
        "deadline_sort": "9999-12-30",
    },
    {
        "title": "Draper Richards Kaplan Foundation — Early-Stage Social Enterprises",
        "url": "https://www.drkfoundation.org/apply/",
        "gender": MULTI_GENDER,
        "highlight": "00B0F0",
        "description": (
            "Rolling philanthropic support (often cited up to about "
            "US$300,000) for early-stage social impact organisations "
            "solving urgent global problems with scalable models. "
            "Relevant if FairBanks/FCHIP presents as a social enterprise "
            "with a clear theory of change for community health "
            "intelligence in underserved peri-urban African settings. "
            "Highly selective; confirm current criteria and apply process "
            "on the official DRK Foundation website before preparing a "
            "full package."
        ),
        "deadline": "Rolling",
        "deadline_sort": "9999-12-31",
    },
]


def word_count(text: str) -> int:
    return len(text.split())


def main() -> None:
    for row in OPPORTUNITIES:
        wc = word_count(row["description"])
        if wc > 100:
            raise SystemExit(f"Description too long ({wc}): {row['title']}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    headers = [
        "Project Title",
        "URL",
        "Brief Description",
        "Application Deadline",
        "Gender Category",
        "Application Status",
        "Submission Status",
        "Submission Date",
        "Application Folder",
    ]
    header_fill = PatternFill("solid", fgColor="0B3D2E")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    alt_fill = PatternFill("solid", fgColor="F3F8F5")
    gender_fill = PatternFill("solid", fgColor="F6E7F2")
    multi_fill = PatternFill("solid", fgColor="E7F0FA")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")

    # Soonest deadline first; unknown / rolling last.
    rows = sorted(
        OPPORTUNITIES,
        key=lambda r: (r["deadline_sort"], r["title"]),
    )

    for i, row in enumerate(rows, 2):
        slug = row.get("app_slug", "")
        highlight = row.get("highlight")
        values = [
            row["title"],
            row["url"],
            row["description"],
            row["deadline"],
            row["gender"],
            "Drafting" if slug else "Not started",
            "unknown",
            "",
            f"applications/{slug}" if slug else "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i, col, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if highlight:
                cell.fill = PatternFill("solid", fgColor=highlight)
            elif i % 2 == 0:
                cell.fill = alt_fill
        gcell = ws.cell(i, 5)
        if not highlight:
            gcell.fill = gender_fill if row["gender"] == GENDER_BASED else multi_fill
        gcell.font = Font(name="Calibri", size=10, bold=True)
        ws.row_dimensions[i].height = 92

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 74
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 22
    ws.row_dimensions[1].height = 26
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"
    ws.freeze_panes = "A2"

    # Scan notes / methodology sheet.
    meta = wb.create_sheet("Scan Notes")
    gb = sum(1 for r in rows if r["gender"] == GENDER_BASED)
    mg = len(rows) - gb
    notes = [
        ("Scanned on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        (
            "Document type",
            "Canonical merged tracker (OFY scan + former peri-urban / FCHIP "
            "complementary scan). Sorted by earliest Application Deadline; "
            "rolling/unknown last. Peri-urban highlight colours preserved.",
        ),
        (
            "Applicant context",
            "FairBanks / FCHIP — a Ugandan community health intelligence "
            "venture (peri-urban Kampala communities; CHW/VHT cascade).",
        ),
        (
            "Sources",
            "Opportunities for Youth plus peri-urban / FCHIP sources "
            "(Grand Challenges Nexa, IDRC ANeSA, Africa Health ExCon, UN Women "
            "SIARP, DIV, FID, EDCTP3, Wellcome Snakebite, Social Shifters, "
            "AFNet, Gilead, DRK, and related).",
        ),
        (
            "Eligibility filter",
            "Included ONLY opportunities that are (a) still open (or "
            "reopening) and (b) open to Ugandan applicants (global, "
            "pan-African / African Union, or Uganda-specific).",
        ),
        (
            "Highlight legend",
            "Green = peri-urban priority women/gender calls. Blue = strong "
            "FCHIP / digital / innovation fits. Red = research awards with "
            "narrower fit. Black = FID portal reopen watch.",
        ),
        (
            "Gender categories",
            f"Gender-based (women/girls): {gb}. Multi-gender (all genders): "
            f"{mg}. Total: {len(rows)}.",
        ),
        (
            "Priority soon",
            "Africa Health-Tech Accelerator (20 Jul), Nexa (22 Jul), AWIEF "
            "(20 Jul), Gadfly (24 Jul), SIARP 2.0 (27 Jul), US State Dept "
            "Uganda SOI / WHS (31 Jul).",
        ),
        (
            "Best fit for FairBanks",
            "US State Dept Uganda Health MOU, Nexa, Africa Health-Tech "
            "Accelerator, AWIEF Pitch n Grow, EDCTP3 DIGIT-02 (consortium), "
            "DIV/FID evidence-to-scale.",
        ),
        (
            "Application packs",
            "Win-win packs (rules + Word/PDF/PPT) live under "
            "applications/{slug}/ where an app_slug is set. Status Drafting "
            "until submitted; peri-urban adds start as Not started.",
        ),
        (
            "Important",
            "ALWAYS confirm deadlines and eligibility on the official "
            "programme page before applying. Never use sites that impersonate "
            "IFC/World Bank grant portals.",
        ),
    ]
    meta["A1"] = "Field"
    meta["B1"] = "Detail"
    meta["A1"].font = Font(bold=True, color="FFFFFF")
    meta["B1"].font = Font(bold=True, color="FFFFFF")
    meta["A1"].fill = header_fill
    meta["B1"].fill = header_fill
    for i, (k, v) in enumerate(notes, 2):
        meta.cell(i, 1, k).font = Font(bold=True)
        meta.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
        meta.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
        meta.row_dimensions[i].height = 42
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 112

    out = Path(os.environ.get("OPPORTUNITIES_OUT", str(DEFAULT_OUT)))
    wb.save(out)
    print(f"Wrote {len(rows)} opportunities to {out}")
    print(f"Gender-based: {gb} | Multi-gender: {mg}")
    for r in rows:
        print(f"- [{r['gender'][:12]:12}] {r['deadline'][:34]:34} | {r['title'][:60]}")


if __name__ == "__main__":
    main()

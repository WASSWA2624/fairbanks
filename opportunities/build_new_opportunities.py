#!/usr/bin/env python3
"""Build opportunities/new_opportunities.xlsx — fresh scan (29 July 2026).

Clears prior rows and writes only still-open, Uganda-eligible FairBanks /
FCHIP fits not already on the working tracker (tracker.xlsx).

Layout per opportunities/rules/source_of_truth.mdc.
Applicant context: FAIRBANKS MEDICAL CENTRE LIMITED — women-led Ugandan
private company (not NGO); Community Reach + FCHIP.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "new_opportunities.xlsx"

GENDER_BASED = "Gender-based (women/girls)"
MULTI_GENDER = "Multi-gender (all genders)"

# Verified open as of 29 July 2026. Prefer official URLs.
# Excludes closed calls, Uganda-blocked countries, Switzerland-only work
# permits, NGO-INGO-only sport RFPs, agriculture-only AI challenges, and
# anything already on opportunities/tracker.xlsx.
OPPORTUNITIES = [
    {
        "title": "Nexa Transition to Scale — Climate × Health (extended)",
        "url": "https://www.grandchallenges.ca/funding-opportunity-nexa/",
        "gender": MULTI_GENDER,
        "description": (
            "Grand Challenges Canada / Science for Africa Foundation "
            "Transition to Scale track for innovations that turn "
            "climate-driven health risk signals into timely care "
            "(mosquito-borne disease, heat, air quality). Funding about "
            "USD 250,000–2,000,000. PoC window closed 22 July; TTS "
            "extended to 12 August 2026, 2 p.m. ET via Fluxx. Uganda "
            "eligible. Strong FCHIP climate–GIS early-warning fit if "
            "MVP evidence is ready to scale. Confirm TTS maturity bar."
        ),
        "deadline": "12 August 2026 (2:00 p.m. ET / 6:00 p.m. UTC)",
        "deadline_sort": "2026-08-12",
        "fit": "FCHIP climate early warning scale",
        "highlight": "00B0F0",
    },
    {
        "title": "CDC — Strengthening Global Health Security via Local Partnerships (Uganda)",
        "url": "https://simpler.grants.gov/opportunity/264249e6-fdbb-4b1c-ac90-23b7d9b07b1b",
        "gender": MULTI_GENDER,
        "description": (
            "CDC cooperative agreement (CDC-RFA-JG-26-0054) to strengthen "
            "Uganda capacities for prevent, detect, and respond — "
            "surveillance, labs, rapid response, and multi-hazard "
            "preparedness. About USD 5M programme total; ~3 awards. "
            "Listing includes for-profit / unrestricted entity types, "
            "so a Ugandan private health company may apply — still a "
            "large USG package. Strong thematic fit with FCHIP "
            "community–facility surveillance intelligence. Confirm "
            "full NOFO eligibility and partnership expectations on "
            "Simpler.Grants.gov before drafting."
        ),
        "deadline": "14 August 2026 (11:59 p.m. ET)",
        "deadline_sort": "2026-08-14",
        "fit": "Uganda health security / surveillance",
        "highlight": "00B0F0",
    },
    {
        "title": "Blue Ambition Fund 2026 (Wilde Ganzen) — Uganda / Ethiopia / South Africa",
        "url": "https://www.wildeganzen.org/programs/blue-ambition-fund/baf-call-for-applications",
        "gender": MULTI_GENDER,
        "description": (
            "Wilde Ganzen Blue Ambition Fund supports non-profits and "
            "social enterprises creating jobs for youth (15–35) and "
            "women via skills-to-employment or support to small growing "
            "businesses. Grants about EUR 100,000–300,000 over 2–3 "
            "years; org annual budget roughly EUR 300k–5M. Uganda "
            "eligible. Aligns with FairBanks livelihoods / CHIS / "
            "IGA empowerment layer more than FCHIP tech. Frame as "
            "social enterprise; confirm budget band and own-contribution "
            "rules in BAF participation requirements."
        ),
        "deadline": "15 August 2026",
        "deadline_sort": "2026-08-15",
        "fit": "Livelihoods / women & youth jobs",
        "highlight": "92D050",
    },
    {
        "title": "Australian High Commission DAP 2026–27 (East Africa — incl. Uganda)",
        "url": "https://kenya.highcommission.gov.au/nair/dap.html",
        "gender": MULTI_GENDER,
        "description": (
            "Direct Aid Program small grants up to AUD 60,000 for "
            "community projects in Burundi, Kenya, Rwanda, Somalia, "
            "Tanzania, or Uganda. 2026–27 priorities include health and "
            "community wellbeing, climate adaptation, gender equality, "
            "and essential services. Apply via SmartyGrants. Typically "
            "targets NGOs/CBOs and registered community groups — "
            "FairBanks would need a clear community project frame "
            "(e.g. outreach / CHW / maternal–child). Confirm applicant "
            "type on the Nairobi DAP page before drafting."
        ),
        "deadline": "16 August 2026 (11:59 p.m. EAT)",
        "deadline_sort": "2026-08-16",
        "fit": "Community health / climate small grant",
    },
    {
        "title": "AU–UNDP YouthConnekt Africa Export Accelerator 2026",
        "url": "https://opportunitiesforyouth.org/2026/07/21/african-union-undp-youthconnekt-africa-export-accelerator-2026-applications-open-for-youth-led-businesses-expanding-across-africa/",
        "gender": MULTI_GENDER,
        "description": (
            "YouthConnekt Africa / UNDP / AU export-readiness accelerator "
            "for youth-led businesses expanding under AfCFTA. Pilot "
            "countries include Uganda. Founders must be under 35 with an "
            "established, export-ready enterprise. Not a health grant; "
            "useful for regional growth of FairBanks / FCHIP services "
            "if founder age and export pathway fit. Confirm official "
            "registration form and sector rules on YouthConnekt / UNDP "
            "pages before applying."
        ),
        "deadline": "16 August 2026",
        "deadline_sort": "2026-08-16",
        "fit": "Youth export growth (age check)",
        "highlight": "FF0000",
    },
    {
        "title": "Africa Deep Tech Challenge 2026 — Laptop LLM (offline AI)",
        "url": "https://africadeeptech.org/challenge-2026/",
        "gender": MULTI_GENDER,
        "description": (
            "Africa Deep Tech Foundation competition to build useful "
            "offline language-model apps on commodity 8 GB laptops "
            "(no cloud). Healthcare / medical support is an allowed "
            "domain — strong match for offline CHW/VHT tools and "
            "clinic decision support under FCHIP. Gate 1 by 25 August "
            "2026 via Devpost; prizes up to USD 8,000 plus residency "
            "and mentorship. Uganda / Africa builders eligible. "
            "Requires a working prototype, not a proposal-only pitch."
        ),
        "deadline": "25 August 2026 (Gate 1)",
        "deadline_sort": "2026-08-25",
        "fit": "Offline AI for CHW / clinic tools",
        "highlight": "00B0F0",
    },
    {
        "title": "Global Health EDCTP3 — DIGIT-02 Digital Innovation & AI in SSA",
        "url": "https://www.edctp.org/edctp-association-as-coordinator-of-global-health-edctp3-projects-2/2026-calls-for-proposals/",
        "gender": MULTI_GENDER,
        "description": (
            "Horizon Europe / Global Health EDCTP3 CSA call "
            "HORIZON-JU-GH-EDCTP3-2026-03-DIGIT-02 to strengthen use of "
            "existing digital health and AI tools for infectious-disease "
            "care and preparedness in sub-Saharan Africa (scale/"
            "interop, not greenfield gadgets). Up to about EUR 2.25M "
            "per project; deadline 2 September 2026, 17:00 Brussels. "
            "Needs a research consortium — FairBanks as community/"
            "facility digital partner. Confirm portal topic page and "
            "consortium lead before joining."
        ),
        "deadline": "2 September 2026 (17:00 Brussels)",
        "deadline_sort": "2026-09-02",
        "fit": "Digital health / AI consortium path",
        "highlight": "00B0F0",
    },
    {
        "title": "Oneness Revival Team (ORT) SEED Grant Program 2026",
        "url": "https://onenessrevivalteam.info/grant-application/",
        "gender": MULTI_GENDER,
        "description": (
            "USD 25,000–50,000 one-year SEED grants aimed at legally "
            "registered grassroots NGOs / CSOs (often with budgets under "
            "USD 100,000). Possible community-reach support only if "
            "FairBanks can partner through an eligible CSO path — "
            "private limited companies are usually out of scope. "
            "Verify authenticity, budget ceiling, and legal-form rules "
            "on the official ORT portal before investing proposal time."
        ),
        "deadline": "20 December 2026",
        "deadline_sort": "2026-12-20",
        "fit": "NGO-only small grant (partner path)",
        "highlight": "FF0000",
    },
]


def word_count(text: str) -> int:
    return len(text.split())


def main() -> None:
    for row in OPPORTUNITIES:
        wc = word_count(row["description"])
        if wc > 100:
            raise SystemExit(f"Description too long ({wc} words): {row['title']}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    headers = [
        "Project Title",
        "URL",
        "Brief Description",
        "Application Deadline",
        "Gender Category",
        "Application Status",
        "Submission Status",
        "Submission Date",
        "Application Folder",
    ]
    header_fill = PatternFill("solid", fgColor="0B3D2E")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    alt_fill = PatternFill("solid", fgColor="F3F8F5")
    gender_fill = PatternFill("solid", fgColor="F6E7F2")
    multi_fill = PatternFill("solid", fgColor="E7F0FA")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")

    rows = sorted(OPPORTUNITIES, key=lambda r: (r["deadline_sort"], r["title"]))

    for i, row in enumerate(rows, 2):
        highlight = row.get("highlight")
        values = [
            row["title"],
            row["url"],
            row["description"],
            row["deadline"],
            row["gender"],
            "Not started",
            "unknown",
            "",
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i, col, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if highlight:
                cell.fill = PatternFill("solid", fgColor=highlight)
            elif i % 2 == 0:
                cell.fill = alt_fill
        gcell = ws.cell(i, 5)
        if not highlight:
            gcell.fill = gender_fill if row["gender"] == GENDER_BASED else multi_fill
        gcell.font = Font(name="Calibri", size=10, bold=True)
        ws.row_dimensions[i].height = 96

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 74
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 20
    ws.row_dimensions[1].height = 26
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"
    ws.freeze_panes = "A2"

    meta = wb.create_sheet("Scan Notes")
    gb = sum(1 for r in rows if r["gender"] == GENDER_BASED)
    mg = len(rows) - gb
    priority = [
        r["title"] for r in rows if r.get("highlight") in {"00B0F0", "92D050"}
    ][:8]
    notes = [
        ("Scanned on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        (
            "Document type",
            "CLEARED and rebuilt new_opportunities.xlsx (29 July 2026). "
            "Fresh open-call scan only. Does not replace "
            "tracker.xlsx. Deduped against the working tracker.",
        ),
        (
            "Applicant context",
            "FAIRBANKS MEDICAL CENTRE LIMITED — Ugandan private company "
            "limited by shares (Co. 80020003843337; TIN 1053370026); "
            "women-led (MD Racheal Nabukeera; 60% women shareholding). "
            "Not an NGO/CBO. Live Community Reach + FCHIP MVP "
            "(CHW/VHT, GIS, climate, EMR/HMS APIs).",
        ),
        (
            "Sources",
            "Opportunities for Youth; www2.fundsforngos.org health/"
            "innovation listings; Grand Challenges Canada Nexa page; "
            "Australian High Commission Nairobi DAP; Wilde Ganzen BAF; "
            "Africa Deep Tech Foundation; EDCTP / Global Health EDCTP3; "
            "Grants.gov CDC Uganda GHS; IDRC/official pages cross-checked.",
        ),
        (
            "Approach",
            "1) Clear prior new-scan rows. 2) Search OFY + fundsforNGOs + "
            "GCC + EDCTP + DAP + Wilde Ganzen for deadlines on/after "
            "29 July 2026. 3) Keep Uganda/Africa/global eligibility. "
            "4) Prefer private-company / women-led / HealthTech / "
            "community-health / climate–health / DPI / livelihoods fits. "
            "5) Prefer official apply URLs. 6) Dedup vs best tracker.",
        ),
        (
            "Eligibility filter",
            "Still open + Ugandans can apply. Flag private-company vs "
            "NGO-only. Exclude country lists without Uganda, "
            "Switzerland work-permit AI fund, INGO-only Adidas Moving "
            "for Change, GC AI Family Planning (Uganda not in country "
            "list), Purpose Earth (capacity closed), French-only AREF "
            "grant-writing cohort, university-only AREF-MRC leadership, "
            "and agriculture-only Innovate Africa Challenge.",
        ),
        (
            "Gender categories",
            f"Gender-based (women/girls): {gb}. Multi-gender: {mg}. "
            f"Total: {len(rows)}.",
        ),
        (
            "Highlight legend",
            "Green = livelihoods / women–youth employment priority. "
            "Blue = strong FCHIP / digital / climate–health / Uganda "
            "health-security fit. Red = narrower or age/partner-path "
            "caveat.",
        ),
        (
            "Priority soon",
            "; ".join(priority)
            if priority
            else "See blue/green highlighted rows sorted by deadline.",
        ),
        (
            "Best FairBanks / FCHIP fits from this scan",
            "1) Nexa TTS (climate early warning scale) 2) Africa Deep "
            "Tech Challenge (offline CHW AI) 3) EDCTP3 DIGIT-02 "
            "(consortium digital health) 4) CDC Uganda GHS "
            "(surveillance — verify entity type) 5) Blue Ambition "
            "(livelihoods / CHIS empowerment layer).",
        ),
        (
            "Excluded (already on working tracker)",
            "AWIEF; Africa Health-Tech Accelerator; Nexa PoC (submitted); "
            "DPI Safeguards (submitted); DoS Uganda MOU; Japan GGP; "
            "Jay Shetty; Yunus; ANeSA LoI; FID reopen watch; DIV; "
            "Gadfly; WHS Youth Group; SIARP; Brandtech; AES Fellowship; "
            "and other titles already listed in tracker.xlsx.",
        ),
        (
            "Excluded (closed / capacity / passed)",
            "ISS Young Changemakers (26 Jul); Feminist Leadership "
            "Accelerator (27 Jul); Purpose Earth 2027 (200-app "
            "capacity closed); Standard Chartered Women in Tech Uganda "
            "(30 Jun); Cartier Women’s Initiative 2027 (16 Jun); "
            "prior new-scan rows past deadline (UPHFP 31 Jul listed "
            "separately — verify if still open on MoH/Makerere pages).",
        ),
        (
            "Excluded (open but unfit / ineligible)",
            "Grand Challenges AI Family Planning (eligible countries "
            "omit Uganda); Prototype Fund Switzerland (Swiss work "
            "permit); Adidas Moving for Change East Africa (INGO + "
            "local S4D only); Innovate Africa Challenge (AI climate-"
            "smart agriculture); AREF French grant-writing cohort; "
            "AREF-MRC Towards Leadership (university/research "
            "employment); UK/US/AU/Canada-only community grants on "
            "fundsforNGOs health page.",
        ),
        (
            "Important",
            "Aggregator pages republish calls — ALWAYS confirm deadline, "
            "eligibility, and apply link on the official funder page. "
            "FairBanks is a private company: many NGO-only calls need a "
            "partner pathway. This file is a scan snapshot, not legal "
            "advice.",
        ),
        ("Build script", "opportunities/build_new_opportunities.py"),
        (
            "Related working tracker",
            "opportunities/tracker.xlsx",
        ),
    ]
    meta["A1"] = "Field"
    meta["B1"] = "Detail"
    meta["A1"].font = Font(bold=True, color="FFFFFF")
    meta["B1"].font = Font(bold=True, color="FFFFFF")
    meta["A1"].fill = header_fill
    meta["B1"].fill = header_fill
    for i, (field, detail) in enumerate(notes, 2):
        meta.cell(i, 1, field).alignment = Alignment(vertical="top", wrap_text=True)
        meta.cell(i, 2, detail).alignment = Alignment(vertical="top", wrap_text=True)
        meta.row_dimensions[i].height = 70
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 96

    fit = wb.create_sheet("Fit Summary")
    fit_headers = ["Deadline sort", "Project Title", "FairBanks fit note", "Gender"]
    for col, h in enumerate(fit_headers, 1):
        cell = fit.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for i, row in enumerate(rows, 2):
        fit.cell(i, 1, row["deadline_sort"])
        fit.cell(i, 2, row["title"])
        fit.cell(i, 3, row.get("fit", ""))
        fit.cell(i, 4, row["gender"])
        for c in range(1, 5):
            fit.cell(i, c).alignment = Alignment(wrap_text=True, vertical="top")
            fit.cell(i, c).border = thin
        fit.row_dimensions[i].height = 36
    fit.column_dimensions["A"].width = 14
    fit.column_dimensions["B"].width = 55
    fit.column_dimensions["C"].width = 36
    fit.column_dimensions["D"].width = 28

    wb.save(OUT)
    print(f"Wrote {OUT} with {len(rows)} opportunities")
    print(f"Gender-based: {gb}, Multi-gender: {mg}")


if __name__ == "__main__":
    main()

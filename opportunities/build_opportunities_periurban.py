#!/usr/bin/env python3
"""Build a NEW FairBanks peri-urban / FCHIP opportunity scan Excel.

Writes opportunities/opportunities_periurban_fchip_2026-07.xlsx.
Does NOT replace or modify opportunities/opportunities.xlsx.

Curated from a deep multi-source scan (18 July 2026) focused on:
community health intelligence, CHW/VHT, digital health, disease prediction,
maternal/child health, peri-urban / underserved African communities,
Uganda-eligible open calls.
"""

from datetime import datetime
from pathlib import Path
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent
DEFAULT_OUT = ROOT / "opportunities_periurban_fchip_2026-07.xlsx"

GENDER_BASED = "Gender-based (women/girls)"
MULTI_GENDER = "Multi-gender (all genders)"

# New scan only — intentionally excludes rows already in opportunities.xlsx.
OPPORTUNITIES = [
    {
        "title": "UN Women — Spotlight Initiative Africa Regional Programme (SIARP) 2.0",
        "url": "https://africa.unwomen.org/en/programme-implementation/2026/06/call-for-proposals-for-civil-society-partnerships-for-implementation-of-the-spotlight-initiative-africa-regional-programme-siarp-20",
        "gender": GENDER_BASED,
        "description": (
            "UN Women call for African civil society partners under SIARP 2.0. "
            "Grants of about US$110,000–548,000 for up to 27 months (from Oct "
            "2026) on ending violence against women and girls, harmful "
            "practices, SRHR, and women- and youth-led movements. Strongest "
            "fit if FairBanks joins a regional consortium linking peri-urban "
            "maternal/adolescent health, CHW outreach, and community "
            "accountability. Open to CSOs, women’s rights and youth-led "
            "groups across Africa, so Ugandan organisations may apply — "
            "confirm regional/multi-country scope on the call page."
        ),
        "deadline": "27 July 2026 (5:00 PM East Africa Time)",
        "deadline_sort": "2026-07-27",
    },
    {
        "title": "AFNet Flexible Grant (African Women for Change Network)",
        "url": "https://afwcnet.org/women-grants",
        "gender": GENDER_BASED,
        "description": (
            "Small flexible grant (about US$5,000) for women leaders and "
            "women-led organisations driving community change in Africa. "
            "Reviewed on a rolling basis. Useful seed support for a "
            "woman-led FairBanks Community Reach or FCHIP pilot activity in "
            "peri-urban Kampala communities (outreach, maternal support, or "
            "CHW coordination). Confirm current application steps and "
            "eligibility on the AFNet / African Women for Change Network "
            "channel before applying; listing details can change."
        ),
        "deadline": "Rolling",
        "deadline_sort": "9999-12-31",
    },
    {
        "title": "Africa Health-Tech Accelerator 2026 (Africa Health Tech / Africa CDC / Health ExCon)",
        "url": "https://accelerator.africahealthexcon.com/apply",
        "gender": MULTI_GENDER,
        "description": (
            "Six-month pan-African accelerator for early-stage health-tech "
            "startups and SMEs. Tracks include digital healthcare, health "
            "data and intelligent technologies, prevention/access services, "
            "and related innovation. Benefits: structured training, "
            "mentorship, seed-funding opportunities, investor Demo Day, and "
            "continental visibility. Needs at least two founders, an MVP, "
            "and a scale-across-Africa commitment. Excellent near-term fit "
            "for FCHIP as a community health intelligence / digital health "
            "venture. Deadline extended to 20 July 2026 — apply immediately."
        ),
        "deadline": "20 July 2026",
        "deadline_sort": "2026-07-20",
    },
    {
        "title": "Nexa Climate and Health Initiative (Grand Challenges Canada / Science for Africa Foundation)",
        "url": "https://www.grandchallenges.ca/rfp-nexa/",
        "gender": MULTI_GENDER,
        "description": (
            "Major open competition for climate–health innovations in "
            "Africa (and LAC for scale). Funds tools that turn climate "
            "risk signals into timely health action on mosquito-borne "
            "disease (e.g. malaria/dengue), extreme heat, and air "
            "quality. Proof of Concept up to US$200,000; Transition to "
            "Scale about US$250,000–2,000,000. PoC applicants must be "
            "incorporated in an eligible African country (Uganda "
            "included). Closest deep-tech match for FCHIP disease early "
            "warning in peri-urban communities. Submit via Fluxx by "
            "22 July 2026, 2:00 p.m. ET."
        ),
        "deadline": "22 July 2026 (2:00 p.m. ET / 6:00 p.m. UTC)",
        "deadline_sort": "2026-07-22",
    },
    {
        "title": "IDRC / GAC / CIHR — ANeSA Cohort 2 Letters of Interest (SRHR in sub-Saharan Africa)",
        "url": "https://idrc-crdi.ca/en/funding/call-letters-interest-addressing-neglected-areas-sexual-and-reproductive-health-and-0",
        "gender": MULTI_GENDER,
        "description": (
            "Open call for Letters of Interest under Addressing Neglected "
            "Areas of SRHR in sub-Saharan Africa (ANeSA). Up to six grants "
            "of up to CAD 1.2 million each (to 36 months) for "
            "gender-transformative implementation research with "
            "under-served populations. Priorities include family planning, "
            "adolescent SRHR, SGBV services, and SRHR advocacy. Led by "
            "SSA-based researchers; Uganda is an eligible country "
            "(additional IDRC approval may apply). Strong research "
            "partnership path for FairBanks peri-urban maternal and "
            "adolescent community health work."
        ),
        "deadline": "23 August 2026 (23:59 ET)",
        "deadline_sort": "2026-08-23",
    },
    {
        "title": "Gilead Global Public Health Awards 2026",
        "url": "https://opportunitiesforyouth.org/2026/05/25/gilead-global-public-health-awards-2026-100000-research-funding-for-early-career-investigators-in-hiv-and-viral-hepatitis/",
        "gender": MULTI_GENDER,
        "description": (
            "Four awards of up to US$100,000 over two years for early-career "
            "investigators in LMICs (including Africa) focused on HIV and "
            "viral hepatitis public health / implementation science. "
            "Applicants need a terminal degree within 10 years, an "
            "in-country mentor, and about 50% research time. Best if a "
            "FairBanks-linked clinician or researcher studies community "
            "prevention, access, or peri-urban care pathways. Confirm on "
            "the official Gilead awards portal before applying."
        ),
        "deadline": "28 August 2026 (11:59 PM CEST)",
        "deadline_sort": "2026-08-28",
    },
    {
        "title": "Social Shifters Global Innovation Challenge 2026",
        "url": "https://www.socialshifters.co/global-innovation-challenge/",
        "gender": MULTI_GENDER,
        "description": (
            "Global challenge for youth-led projects and startups aligned "
            "with at least one SDG, offering up to about US$15,000 in "
            "support. Suitable for a young FairBanks / FCHIP team member "
            "piloting a community health, digital CHW, or peri-urban "
            "prevention idea. Lower funding than major grants but useful "
            "for visibility and early traction. Deadline is 31 August 2026, "
            "5pm UTC. Confirm age rules and award tracks on the Social "
            "Shifters challenge page before investing proposal time."
        ),
        "deadline": "31 August 2026 (5:00 PM UTC)",
        "deadline_sort": "2026-08-31",
    },
    {
        "title": "Fund for Innovation in Development (FID) — France (platform reopens 1 Sep 2026)",
        "url": "https://fundinnovation.dev/en/launch-project",
        "gender": MULTI_GENDER,
        "description": (
            "French open innovation fund for solutions reducing poverty "
            "and inequality in ODA-eligible countries (Uganda eligible). "
            "Health is a priority sector alongside education, climate, and "
            "gender. Staged grants from pilot to scale (historically up to "
            "about €4 million at scale). Portal is closed for a technical "
            "upgrade and reopens 1 September 2026 for most stages "
            "(Preparation grants later). Strong longer-horizon fit for "
            "evidence-backed FCHIP pilots in peri-urban communities."
        ),
        "deadline": "Reopens 1 September 2026 — then rolling",
        "deadline_sort": "2026-09-01",
    },
    {
        "title": "DIV Fund (Development Innovation Ventures) — Open Innovation Grants",
        "url": "https://www.div.fund/apply",
        "gender": MULTI_GENDER,
        "description": (
            "Evidence-driven open innovation fund accepting applications "
            "year-round. Stages roughly: Pilot up to ~US$200,000; Testing "
            "higher; Scale up to about US$1.5 million. Funds innovations "
            "that can cost-effectively improve millions of lives in "
            "low- and middle-income countries. Open to nonprofits, "
            "startups, universities, and social enterprises — Ugandan "
            "applicants eligible. Good medium-term pathway for FCHIP if "
            "you can show a clear pilot plan, cost-effectiveness case, "
            "and path to scale across peri-urban and district health "
            "systems."
        ),
        "deadline": "Rolling (applications accepted year-round)",
        "deadline_sort": "9999-12-30",
    },
    {
        "title": "Draper Richards Kaplan Foundation — Early-Stage Social Enterprises",
        "url": "https://www.drkfoundation.org/apply/",
        "gender": MULTI_GENDER,
        "description": (
            "Rolling philanthropic support (often cited up to about "
            "US$300,000) for early-stage social impact organisations "
            "solving urgent global problems with scalable models. "
            "Relevant if FairBanks/FCHIP presents as a social enterprise "
            "with a clear theory of change for community health "
            "intelligence in underserved peri-urban African settings. "
            "Highly selective; confirm current criteria and apply process "
            "on the official DRK Foundation website before preparing a "
            "full package."
        ),
        "deadline": "Rolling",
        "deadline_sort": "9999-12-31",
    },
]


def word_count(text: str) -> int:
    return len(text.split())


def main() -> None:
    for row in OPPORTUNITIES:
        wc = word_count(row["description"])
        if wc > 100:
            raise SystemExit(f"Description too long ({wc}): {row['title']}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    headers = [
        "Project Title",
        "URL",
        "Brief Description",
        "Application Deadline",
        "Gender Category",
        "Application Status",
        "Submission Status",
        "Submission Date",
        "Application Folder",
    ]
    header_fill = PatternFill("solid", fgColor="0B3D2E")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    alt_fill = PatternFill("solid", fgColor="F3F8F5")
    gender_fill = PatternFill("solid", fgColor="F6E7F2")
    multi_fill = PatternFill("solid", fgColor="E7F0FA")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")

    rows = sorted(
        OPPORTUNITIES,
        key=lambda r: (0 if r["gender"] == GENDER_BASED else 1, r["deadline_sort"], r["title"]),
    )

    for i, row in enumerate(rows, 2):
        values = [
            row["title"],
            row["url"],
            row["description"],
            row["deadline"],
            row["gender"],
            "Not started",
            "unknown",
            "",
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i, col, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if i % 2 == 0:
                cell.fill = alt_fill
        gcell = ws.cell(i, 5)
        gcell.fill = gender_fill if row["gender"] == GENDER_BASED else multi_fill
        gcell.font = Font(name="Calibri", size=10, bold=True)
        ws.row_dimensions[i].height = 100

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 74
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 18
    ws.row_dimensions[1].height = 26
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"
    ws.freeze_panes = "A2"

    meta = wb.create_sheet("Scan Notes")
    gb = sum(1 for r in rows if r["gender"] == GENDER_BASED)
    mg = len(rows) - gb
    notes = [
        ("Document type", "NEW peri-urban / FCHIP funder scan. Does NOT replace opportunities/opportunities.xlsx."),
        ("Scanned on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        (
            "Applicant context",
            "FairBanks / FCHIP — Ugandan community health intelligence venture; peri-urban Kampala communities "
            "(e.g. Bukoto, Kyebando, Kisaasi, Kamwokya, Kikaaya); CHW/VHT cascade; maternal/child & NCD screening.",
        ),
        (
            "Fit filter",
            "Community health intelligence, CHW/VHT tools, digital health, disease early warning, maternal/child "
            "health, peri-urban / underserved African settings, AI-for-health, health systems / SRHR research.",
        ),
        (
            "Sources scanned",
            "Grand Challenges Canada (Nexa), Science for Africa Foundation, IDRC/CIHR/GAC (ANeSA), Africa Health "
            "ExCon Accelerator, UN Women Africa (SIARP 2.0), DIV Fund, FID France, Opportunities for Youth, "
            "fundsforNGOs, africanngos.org July–Aug 2026 roundup, Global Innovation Fund, GSMA Innovation Fund, "
            "Villgro Africa, CARE SheConnects, reach52, Luena Foundation, la Caixa Child Survival, EVAH/J-PAL, "
            "AJA Foundation, Echoing Green, AUDA-NEPAD HGS.",
        ),
        (
            "Eligibility filter",
            "Included only still-open (or reopening soon) opportunities open to Ugandan applicants "
            "(global, pan-African, East Africa, or Uganda). Closed or country-restricted calls noted below.",
        ),
        (
            "Excluded (closed / paused)",
            "SheConnects CARE Digital Accelerator (8 Jul), Inclusive AI HAIDI (17 Jul), SAFEStart+ Uganda "
            "(30 Jun), Luena Eneko Medical Grants (31 May), la Caixa Child Survival (26 May), UNICEF Climate "
            "Ventures (17 May), AUDA-NEPAD HGS Accelerator (19 May), GSMA Green Transition (6 Apr), EVAH "
            "Spring RFP (1 Apr), Gates Cost-Disruptive Diagnostics (28 Apr), reach52 Growth Partner Q1 "
            "(31 Mar), Echoing Green 2026 Fellowship (closed), AJA Foundation new grants paused, GIF not "
            "accepting applications.",
        ),
        (
            "Excluded (weak / wrong fit)",
            "Darwin biodiversity, cultural heritage funds, country-only non-Uganda rights calls, pure "
            "sport-for-development without health intelligence link, Nigeria/Ethiopia-only CSO funds.",
        ),
        (
            "Gender categories",
            f"Gender-based (women/girls): {gb}. Multi-gender (all genders): {mg}. Total: {len(rows)}.",
        ),
        (
            "Priority soon (peri-urban / FCHIP)",
            "1) Africa Health-Tech Accelerator (20 Jul) 2) Nexa climate–health early warning (22 Jul) "
            "3) UN Women SIARP 2.0 if regional consortium ready (27 Jul) 4) IDRC ANeSA LoI (23 Aug) "
            "5) DIV Fund rolling evidence pathway 6) FID portal reopen 1 Sep.",
        ),
        (
            "Best peri-urban fits",
            "Nexa (mosquito-borne early warning for dense peri-urban settlements), Africa Health-Tech "
            "Accelerator (FCHIP digital health MVP), DIV/FID (scale path), ANeSA (peri-urban adolescent/"
            "maternal SRHR research with university partner).",
        ),
        (
            "Important",
            "Deadlines and eligibility were checked against official or aggregator pages on the scan date. "
            "ALWAYS re-confirm on the official programme page before applying — details change quickly.",
        ),
        (
            "Relation to main tracker",
            "Keep opportunities.xlsx as the primary active tracker. This file is a complementary deep scan "
            "focused on peri-urban + FCHIP funders discovered outside the first Opportunities for Youth pass.",
        ),
    ]
    meta["A1"] = "Field"
    meta["B1"] = "Detail"
    meta["A1"].font = Font(bold=True, color="FFFFFF")
    meta["B1"].font = Font(bold=True, color="FFFFFF")
    meta["A1"].fill = header_fill
    meta["B1"].fill = header_fill
    for i, (k, v) in enumerate(notes, 2):
        meta.cell(i, 1, k).font = Font(bold=True)
        meta.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
        meta.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
        meta.row_dimensions[i].height = 48
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 118

    out = Path(os.environ.get("OPPORTUNITIES_PERIURBAN_OUT", str(DEFAULT_OUT)))
    wb.save(out)
    print(f"Wrote {len(rows)} opportunities to {out}")
    print(f"Gender-based: {gb} | Multi-gender: {mg}")
    print("NOTE: opportunities.xlsx was NOT modified.")
    for r in rows:
        print(f"- [{r['gender'][:12]:12}] {r['deadline'][:34]:34} | {r['title'][:60]}")


if __name__ == "__main__":
    main()
